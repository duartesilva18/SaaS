from fastapi import APIRouter, Request, HTTPException, Depends, Header
from sqlalchemy.orm import Session
from sqlalchemy import func, case
import requests
import json
import logging
import re
import hmac
import hashlib
import uuid
from collections import defaultdict
from datetime import datetime, timedelta, date, timezone
from typing import Optional, Dict, List
from concurrent.futures import ThreadPoolExecutor
from types import SimpleNamespace
import io
import tempfile
import time
import unicodedata
from difflib import SequenceMatcher

from ..core.config import settings
from ..core.dependencies import get_db, SessionLocal
from ..models import database as models
from ..core.limiter import limiter
from ..core.telegram_translations import get_telegram_t

logger = logging.getLogger("telegram_webhook")

# Limite da API Telegram: 4096 caracteres por mensagem
TELEGRAM_MAX_MESSAGE_LENGTH = 4090

# Memória de conversa: últimas mensagens por chat_id (max 5, TTL ~30min)
_conversation_memory: Dict[str, List[Dict]] = defaultdict(list)
_CONVERSATION_MAX = 5
_CONVERSATION_TTL = 1800  # 30 min


def _add_to_memory(chat_id: str, role: str, content: str):
    """Adiciona mensagem à memória de conversa."""
    now = time.time()
    mem = _conversation_memory[chat_id]
    mem.append({"role": role, "content": content[:500], "ts": now})
    # Limpar antigas
    _conversation_memory[chat_id] = [m for m in mem if now - m["ts"] < _CONVERSATION_TTL][-_CONVERSATION_MAX:]


def _get_memory(chat_id: str) -> List[Dict]:
    """Devolve histórico de conversa recente (role, content)."""
    now = time.time()
    mem = _conversation_memory.get(chat_id, [])
    return [{"role": m["role"], "content": m["content"]} for m in mem if now - m["ts"] < _CONVERSATION_TTL]


class AIUnavailableError(Exception):
    """Levantada quando a IA (OpenAI) não está disponível (quota/plano esgotado)."""
    pass


# Mapeamento emoji -> nome de categoria (PT e EN)
EMOJI_CATEGORY_MAP = {
    '🍕': ['Alimentação', 'Food', 'Restaurantes', 'Restaurants'],
    '🍔': ['Alimentação', 'Food', 'Restaurantes', 'Restaurants'],
    '🍽️': ['Alimentação', 'Food', 'Restaurantes', 'Restaurants'],
    '☕': ['Alimentação', 'Food', 'Cafés', 'Coffee'],
    '🍺': ['Alimentação', 'Food', 'Bares', 'Bars'],
    '🍷': ['Alimentação', 'Food', 'Bares', 'Bars'],
    '⛽': ['Transportes', 'Transport', 'Gasolina', 'Fuel'],
    '🚗': ['Transportes', 'Transport'],
    '🚕': ['Transportes', 'Transport', 'Uber', 'Taxi'],
    '🚌': ['Transportes', 'Transport'],
    '✈️': ['Viagens', 'Travel'],
    '🏨': ['Viagens', 'Travel', 'Alojamento', 'Accommodation'],
    '🏥': ['Saúde', 'Health'],
    '💊': ['Saúde', 'Health', 'Farmácia', 'Pharmacy'],
    '🏠': ['Habitação', 'Housing', 'Renda', 'Rent'],
    '💡': ['Utilidades', 'Utilities', 'Eletricidade', 'Electricity'],
    '📱': ['Tecnologia', 'Technology', 'Telecomunicações', 'Telecom'],
    '💻': ['Tecnologia', 'Technology'],
    '🎮': ['Entretenimento', 'Entertainment', 'Jogos', 'Games'],
    '🎬': ['Entretenimento', 'Entertainment', 'Cinema'],
    '🎵': ['Entretenimento', 'Entertainment', 'Música', 'Music'],
    '👕': ['Vestuário', 'Clothing', 'Roupa', 'Clothes'],
    '👟': ['Vestuário', 'Clothing', 'Calçado', 'Shoes'],
    '🎓': ['Educação', 'Education'],
    '📚': ['Educação', 'Education', 'Livros', 'Books'],
    '💰': ['Receitas', 'Income', 'Salário', 'Salary'],
    '🏋️': ['Desporto', 'Sports', 'Ginásio', 'Gym'],
    '💇': ['Cuidados pessoais', 'Personal care'],
    '🐕': ['Animais', 'Pets'],
    '🐈': ['Animais', 'Pets'],
    '🎁': ['Presentes', 'Gifts'],
    '🛒': ['Supermercado', 'Groceries', 'Compras', 'Shopping'],
    '🧹': ['Casa', 'Home', 'Limpeza', 'Cleaning'],
}


def _match_emoji_to_category(text: str, categories, language: str = 'pt'):
    """Se o texto começa com um emoji mapeado, tenta encontrar a categoria correspondente."""
    for emoji, cat_names in EMOJI_CATEGORY_MAP.items():
        if emoji in text:
            for cat in categories:
                cat_lower = (cat.name or "").lower()
                for cn in cat_names:
                    if cn.lower() == cat_lower or cn.lower() in cat_lower:
                        return cat, emoji
    return None, None


def _check_budget_alerts(workspace_id, category_id, db, t) -> Optional[str]:
    """Verifica se uma transação ultrapassou ou está perto do limite de orçamento de uma categoria."""
    cat = db.query(models.Category).filter(models.Category.id == category_id).first()
    if not cat or not cat.monthly_limit_cents or cat.monthly_limit_cents <= 0:
        return None
    today = date.today()
    first_day = today.replace(day=1)
    spent_cents = db.query(
        func.sum(func.abs(models.Transaction.amount_cents))
    ).filter(
        models.Transaction.workspace_id == workspace_id,
        models.Transaction.category_id == category_id,
        models.Transaction.amount_cents < 0,
        models.Transaction.transaction_date >= first_day,
        models.Transaction.transaction_date <= today,
    ).scalar() or 0
    spent = int(spent_cents) / 100
    limit_val = cat.monthly_limit_cents / 100
    percent = int((spent / limit_val) * 100) if limit_val > 0 else 0
    if percent >= 100:
        over = spent - limit_val
        return t('budget_alert_exceeded').format(
            category=cat.name, spent=f"{spent:.2f}", limit=f"{limit_val:.2f}",
            percent=percent, over=f"{over:.2f}",
        )
    elif percent >= 80:
        remaining = limit_val - spent
        return t('budget_alert_warning').format(
            category=cat.name, spent=f"{spent:.2f}", limit=f"{limit_val:.2f}",
            percent=percent, remaining=f"{remaining:.2f}",
        )
    return None


def _check_streak(workspace_id, db, t) -> Optional[str]:
    """Verifica se o user tem uma streak de dias seguidos a registar transações."""
    today = date.today()
    streak = 0
    for i in range(365):
        check_date = today - timedelta(days=i)
        has_tx = db.query(models.Transaction.id).filter(
            models.Transaction.workspace_id == workspace_id,
            models.Transaction.transaction_date == check_date,
        ).first()
        if has_tx:
            streak += 1
        else:
            break
    if streak == 0:
        return None
    if streak == 100:
        return t('streak_milestone_100')
    if streak == 30:
        return t('streak_milestone_30')
    if streak == 7:
        return t('streak_milestone_7')
    if streak >= 3 and streak % 5 == 0:
        return t('streak_message').format(days=streak)
    return None


def _check_month_comparison(workspace_id, db, t, language='pt') -> Optional[str]:
    """No início do mês (dia 1-3), envia comparação com o mês anterior se ainda não enviou."""
    today = date.today()
    if today.day > 3:
        return None
    first_day = today.replace(day=1)
    curr_count = db.query(func.count(models.Transaction.id)).filter(
        models.Transaction.workspace_id == workspace_id,
        models.Transaction.transaction_date >= first_day,
    ).scalar() or 0
    if curr_count > 1:
        return None
    if first_day.month == 1:
        prev_first = first_day.replace(year=first_day.year - 1, month=12)
    else:
        prev_first = first_day.replace(month=first_day.month - 1)
    prev_last = first_day - timedelta(days=1)
    prev_expenses = abs(int(db.query(
        func.sum(case((models.Transaction.amount_cents < 0, models.Transaction.amount_cents), else_=0))
    ).filter(
        models.Transaction.workspace_id == workspace_id,
        models.Transaction.transaction_date >= prev_first,
        models.Transaction.transaction_date <= prev_last,
    ).scalar() or 0)) / 100
    if prev_first.month == 1:
        prev2_first = prev_first.replace(year=prev_first.year - 1, month=12)
    else:
        prev2_first = prev_first.replace(month=prev_first.month - 1)
    prev2_last = prev_first - timedelta(days=1)
    prev2_expenses = abs(int(db.query(
        func.sum(case((models.Transaction.amount_cents < 0, models.Transaction.amount_cents), else_=0))
    ).filter(
        models.Transaction.workspace_id == workspace_id,
        models.Transaction.transaction_date >= prev2_first,
        models.Transaction.transaction_date <= prev2_last,
    ).scalar() or 0)) / 100
    if prev_expenses == 0 and prev2_expenses == 0:
        return None
    month_names_pt = {1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'}
    month_names_en = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}
    names = month_names_pt if language == 'pt' else month_names_en
    prev_name = names.get(prev_first.month, str(prev_first.month))
    prev2_name = names.get(prev2_first.month, str(prev2_first.month))
    if prev2_expenses > 0:
        diff = abs(prev_expenses - prev2_expenses)
        percent = int((diff / prev2_expenses) * 100)
        if prev_expenses < prev2_expenses:
            trend = t('month_trend_better').format(diff=f"{diff:.2f}", percent=percent)
        elif prev_expenses > prev2_expenses:
            trend = t('month_trend_worse').format(diff=f"{diff:.2f}", percent=percent)
        else:
            trend = t('month_trend_equal')
    else:
        trend = ""
    return t('month_comparison').format(
        prev_month=prev_name, prev_expenses=f"{prev_expenses:.2f}",
        curr_month=prev2_name, curr_expenses=f"{prev2_expenses:.2f}",
        trend=trend,
    )


def _generate_insight(workspace_id, category_id, db, t) -> Optional[str]:
    """Gera um insight comparando gastos da categoria com o mês anterior."""
    cat = db.query(models.Category).filter(models.Category.id == category_id).first()
    if not cat or cat.type != 'expense':
        return None
    today = date.today()
    first_day = today.replace(day=1)
    # Current month
    curr = db.query(
        func.sum(func.abs(models.Transaction.amount_cents))
    ).filter(
        models.Transaction.workspace_id == workspace_id,
        models.Transaction.category_id == category_id,
        models.Transaction.amount_cents < 0,
        models.Transaction.transaction_date >= first_day,
    ).scalar() or 0
    # Previous month
    if first_day.month == 1:
        prev_first = first_day.replace(year=first_day.year - 1, month=12)
    else:
        prev_first = first_day.replace(month=first_day.month - 1)
    prev_last = first_day - timedelta(days=1)
    prev = db.query(
        func.sum(func.abs(models.Transaction.amount_cents))
    ).filter(
        models.Transaction.workspace_id == workspace_id,
        models.Transaction.category_id == category_id,
        models.Transaction.amount_cents < 0,
        models.Transaction.transaction_date >= prev_first,
        models.Transaction.transaction_date <= prev_last,
    ).scalar() or 0
    if prev == 0:
        return None
    percent_change = int(((int(curr) - int(prev)) / int(prev)) * 100)
    if percent_change > 20:
        return t('insight_spending_up').format(percent=percent_change, category=cat.name)
    elif percent_change < -20:
        return t('insight_spending_down').format(percent=abs(percent_change), category=cat.name)
    return None


class OpenAIRateLimitError(Exception):
    """Levantada quando a OpenAI devolve 429 Too Many Requests (limite de pedidos excedido)."""
    pass


def _telegram_lang(from_user: Optional[dict]) -> str:
    """Infer bot language from Telegram user (when app user has no language set)."""
    code = (from_user or {}).get("language_code") or "pt"
    return "en" if (code and code.lower().startswith("en")) else "pt"


def _html_escape(s: str) -> str:
    """Escape for Telegram HTML parse_mode (evita que descrições quebrem a mensagem)."""
    if not s:
        return ""
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _shorten_description_for_list(description: str, max_len: int = 45) -> str:
    """
    Para a lista de transações no Telegram: descrições muito longas ou texto de extrato
    (Titular, ContaPT, Saldo disponível, Movimentos sua conta) são reduzidas para a mensagem ficar legível.
    """
    if not description or not description.strip():
        return "—"
    d = description.strip()
    # Texto típico de extrato bancário → label curta
    bank_keywords = (
        "titular", "contapt", "saldo disponível", "movimentos sua conta",
        "movimento", "movimentos", "conta ", "saldo"
    )
    d_lower = d.lower()
    if any(k in d_lower for k in bank_keywords) and len(d) > 40:
        return "Movimento conta"
    if len(d) <= max_len:
        return d
    return d[: max_len - 1].rstrip() + "…"


def _origin_line(inference_source: Optional[str], t) -> str:
    """Returns a short origin label for category (e.g. 'Por cache') or empty string."""
    if not inference_source:
        return ""
    src = (inference_source or "").lower()
    if "cache" in src or src == "cache_telegram":
        key = "source_cache"
    elif "history" in src or "similar" in src:
        key = "source_history"
    elif "openai" in src or "vision" in src:
        key = "source_openai"
    elif "explicit" in src:
        key = "source_explicit"
    else:
        key = "source_fallback"
    return t("origin_suffix", origin=t(key))


def _date_line(transaction_date: date, t) -> str:
    """Returns date line for message if date is not today, else empty."""
    if transaction_date == date.today():
        return ""
    return t("date_line", date=transaction_date.strftime("%d/%m/%Y"))

router = APIRouter(prefix='/telegram', tags=['webhooks'])

# Rate Limiting
_rate_limit_store = defaultdict(list)  # chat_id -> [timestamps]
_rate_limit_window = timedelta(minutes=1)
_rate_limit_max_messages = 10  # Máximo 10 mensagens por minuto

# Idempotência: update_id já processados (TTL ~5 min)
_processed_updates: Dict[int, datetime] = {}
_processed_updates_ttl = timedelta(minutes=5)
PENDING_STALE_HOURS = 24


def _is_duplicate_update(update_id: int) -> bool:
    """True se este update_id já foi processado (evitar duplicados)."""
    if update_id is None:
        return False
    now = datetime.now()
    for uid, ts in list(_processed_updates.items()):
        if now - ts > _processed_updates_ttl:
            del _processed_updates[uid]
    if update_id in _processed_updates:
        return True
    _processed_updates[update_id] = now
    return False

def check_rate_limit(chat_id: str) -> bool:
    """Verifica se o chat_id está dentro do limite de rate"""
    now = datetime.now()
    # Limpar timestamps antigos
    _rate_limit_store[chat_id] = [
        ts for ts in _rate_limit_store[chat_id]
        if now - ts < _rate_limit_window
    ]
    
    # Verificar limite
    if len(_rate_limit_store[chat_id]) >= _rate_limit_max_messages:
        return False  # Limite excedido
    
    _rate_limit_store[chat_id].append(now)
    return True

# ==================== AI ROUTER: GPT-4o-mini conversational assistant ====================

# Rate limit separado para chamadas GPT (5/min por user)
_gpt_rate_limit_store: Dict[str, list] = defaultdict(list)
_gpt_rate_limit_window = timedelta(minutes=1)
_gpt_rate_limit_max = 5


def _check_gpt_rate_limit(chat_id: str) -> bool:
    """True se o user pode fazer mais chamadas GPT neste minuto."""
    now = datetime.now()
    _gpt_rate_limit_store[chat_id] = [
        ts for ts in _gpt_rate_limit_store[chat_id]
        if now - ts < _gpt_rate_limit_window
    ]
    if len(_gpt_rate_limit_store[chat_id]) >= _gpt_rate_limit_max:
        return False
    _gpt_rate_limit_store[chat_id].append(now)
    return True


def _is_obvious_transaction(text: str) -> bool:
    """
    Fast-path: se a mensagem claramente contem um valor monetario, e uma transacao.
    Nestes casos saltamos o GPT e vamos direto para parse_transaction (mais rapido).
    """
    # Padroes obvios: "15€", "15 euros", "15e", "15,50€", numeros seguidos de moeda
    return bool(re.search(
        r'\d+(?:[.,]\d+)?\s*(?:€|eur(?:os?)?|e)\b',
        text,
        re.IGNORECASE,
    ))


def _build_financial_context(user: models.User, workspace: models.Workspace, db: Session) -> str:
    """
    Constroi um resumo financeiro compacto para o system prompt do GPT.
    Inclui: categorias, totais do mes, top gastos. Nao inclui dados pessoais.
    """
    today = date.today()
    first_day = today.replace(day=1)

    # Categorias do workspace
    categories = db.query(models.Category).filter(
        models.Category.workspace_id == workspace.id
    ).all()
    expense_cats = [c.name for c in categories if c.type == 'expense']
    income_cats = [c.name for c in categories if c.type == 'income']

    # Totais do mes
    # Nota: SQLAlchemy 2.x exige que os "whens" sejam passados como argumentos posicionais,
    # não como lista. Usamos tuplos individuais em vez de lista de tuplos.
    q = db.query(
        func.sum(case((models.Transaction.amount_cents < 0, models.Transaction.amount_cents), else_=0)),
        func.sum(case((models.Transaction.amount_cents > 0, models.Transaction.amount_cents), else_=0)),
        func.count(models.Transaction.id),
    ).filter(
        models.Transaction.workspace_id == workspace.id,
        models.Transaction.transaction_date >= first_day,
        models.Transaction.transaction_date <= today,
    ).first()
    expenses_cents = abs(int(q[0] or 0))
    income_cents = int(q[1] or 0)
    tx_count = int(q[2] or 0)
    balance_cents = income_cents - expenses_cents

    # Top 5 categorias de despesa do mes
    top_cats = db.query(
        models.Category.name,
        func.sum(func.abs(models.Transaction.amount_cents)).label('total'),
    ).join(
        models.Transaction, models.Transaction.category_id == models.Category.id
    ).filter(
        models.Transaction.workspace_id == workspace.id,
        models.Transaction.transaction_date >= first_day,
        models.Transaction.amount_cents < 0,
    ).group_by(models.Category.name).order_by(func.sum(func.abs(models.Transaction.amount_cents)).desc()).limit(5).all()

    top_str = ", ".join(f"{name} {int(total)/100:.0f}€" for name, total in top_cats) if top_cats else "sem dados"

    currency = getattr(user, 'currency', 'EUR') or 'EUR'

    return (
        f"Hoje: {today.strftime('%d/%m/%Y')}. Moeda: {currency}.\n"
        f"CATEGORIAS DE DESPESA: {', '.join(expense_cats[:25]) if expense_cats else 'nenhuma'}\n"
        f"CATEGORIAS DE RECEITA: {', '.join(income_cats[:15]) if income_cats else 'nenhuma'}\n"
        f"RESUMO DO MES ({today.strftime('%B %Y')}):\n"
        f"  Despesas: {expenses_cents/100:.2f}€ | Receitas: {income_cents/100:.2f}€ | Saldo: {balance_cents/100:.2f}€ | {tx_count} transacoes\n"
        f"  Top gastos: {top_str}\n"
    )


def _query_financial_data(query_type: str, workspace: models.Workspace, db: Session, params: dict) -> str:
    """
    Executa queries na BD com base no intent do GPT.
    Retorna texto formatado com os resultados para o GPT montar a resposta final.
    """
    today = date.today()
    first_day = today.replace(day=1)

    if query_type == "gastos_categoria":
        cat_name = (params.get("category") or "").strip()
        if not cat_name:
            return "Nenhuma categoria especificada."
        cat = db.query(models.Category).filter(
            models.Category.workspace_id == workspace.id,
            func.lower(models.Category.name).contains(cat_name.lower()),
        ).first()
        if not cat:
            return f"Categoria '{cat_name}' nao encontrada."
        period_start = params.get("period_start")
        period_end = params.get("period_end")
        if period_start:
            try:
                start = datetime.strptime(period_start, "%Y-%m-%d").date()
            except Exception:
                start = first_day
        else:
            start = first_day
        if period_end:
            try:
                end = datetime.strptime(period_end, "%Y-%m-%d").date()
            except Exception:
                end = today
        else:
            end = today
        txs = db.query(models.Transaction).filter(
            models.Transaction.workspace_id == workspace.id,
            models.Transaction.category_id == cat.id,
            models.Transaction.transaction_date >= start,
            models.Transaction.transaction_date <= end,
        ).order_by(models.Transaction.amount_cents).limit(20).all()
        total = sum(abs(t.amount_cents) for t in txs)
        lines = [f"- {t.description}: {abs(t.amount_cents)/100:.2f}€ ({t.transaction_date.strftime('%d/%m')})" for t in txs[:10]]
        return (
            f"Categoria: {cat.name}\nPeriodo: {start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}\n"
            f"Total: {total/100:.2f}€ ({len(txs)} transacoes)\n"
            f"Transacoes:\n" + "\n".join(lines)
        )

    elif query_type == "total_periodo":
        period_start = params.get("period_start")
        period_end = params.get("period_end")
        if period_start:
            try:
                start = datetime.strptime(period_start, "%Y-%m-%d").date()
            except Exception:
                start = first_day
        else:
            start = first_day
        if period_end:
            try:
                end = datetime.strptime(period_end, "%Y-%m-%d").date()
            except Exception:
                end = today
        else:
            end = today

        q = db.query(
            func.sum(case((models.Transaction.amount_cents < 0, models.Transaction.amount_cents), else_=0)),
            func.sum(case((models.Transaction.amount_cents > 0, models.Transaction.amount_cents), else_=0)),
            func.count(models.Transaction.id),
        ).filter(
            models.Transaction.workspace_id == workspace.id,
            models.Transaction.transaction_date >= start,
            models.Transaction.transaction_date <= end,
        ).first()
        expenses = abs(int(q[0] or 0))
        income = int(q[1] or 0)
        count = int(q[2] or 0)
        return (
            f"Periodo: {start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}\n"
            f"Despesas: {expenses/100:.2f}€ | Receitas: {income/100:.2f}€ | Saldo: {(income-expenses)/100:.2f}€ | {count} transacoes"
        )

    elif query_type == "top_despesas":
        limit = min(int(params.get("limit", 10)), 20)
        period_start = params.get("period_start")
        if period_start:
            try:
                start = datetime.strptime(period_start, "%Y-%m-%d").date()
            except Exception:
                start = first_day
        else:
            start = first_day
        txs = db.query(models.Transaction).join(
            models.Category, models.Transaction.category_id == models.Category.id
        ).filter(
            models.Transaction.workspace_id == workspace.id,
            models.Transaction.transaction_date >= start,
            models.Transaction.amount_cents < 0,
        ).order_by(models.Transaction.amount_cents).limit(limit).all()
        lines = []
        for i, t in enumerate(txs, 1):
            cat = db.query(models.Category).filter(models.Category.id == t.category_id).first()
            cat_name = cat.name if cat else "?"
            lines.append(f"{i}. {t.description} — {abs(t.amount_cents)/100:.2f}€ ({cat_name}, {t.transaction_date.strftime('%d/%m')})")
        return f"Top {len(txs)} maiores despesas desde {start.strftime('%d/%m/%Y')}:\n" + "\n".join(lines)

    elif query_type == "por_categoria":
        start = first_day
        cats = db.query(
            models.Category.name,
            func.sum(func.abs(models.Transaction.amount_cents)).label('total'),
            func.count(models.Transaction.id).label('cnt'),
        ).join(
            models.Transaction, models.Transaction.category_id == models.Category.id
        ).filter(
            models.Transaction.workspace_id == workspace.id,
            models.Transaction.transaction_date >= start,
            models.Transaction.amount_cents < 0,
        ).group_by(models.Category.name).order_by(func.sum(func.abs(models.Transaction.amount_cents)).desc()).all()
        lines = [f"- {name}: {int(total)/100:.2f}€ ({cnt} tx)" for name, total, cnt in cats]
        return f"Despesas por categoria ({today.strftime('%B %Y')}):\n" + "\n".join(lines) if lines else "Sem despesas este mes."

    return "Tipo de consulta desconhecido."


def _build_system_prompt(context: str, language: str) -> str:
    """Constroi o system prompt para o AI Router."""
    lang_instruction = "Responde sempre em português." if language == "pt" else "Always reply in English."

    return f"""Tu es o Finly, um assistente financeiro pessoal no Telegram. Personalidade: calmo, zen, direto, usa emojis com moderacao.
{lang_instruction}

{context}

INSTRUCOES RIGOROSAS — responde SEMPRE em JSON valido, sem texto extra:

1. TRANSACAO — se a mensagem regista um gasto/receita (contem valor, descricao de compra, pagamento, etc.):
{{"intent":"transaction","transactions":[{{"amount":15.0,"description":"Almoco","type":"expense","category":"Alimentacao","date":"YYYY-MM-DD"}}]}}
- amount: valor positivo (float)
- type: "expense" ou "income"
- category: nome EXATO de uma categoria existente (lista acima). Prefere SEMPRE uma categoria especifica. Usa "Despesas gerais"/"General expenses" APENAS como ultimo recurso quando nenhuma outra se aplica.
- date: se a mensagem menciona uma data (ontem, dia 15, 28/01, janeiro, yesterday, etc.), converte para "YYYY-MM-DD". Se nao menciona data, usa null.
- Para multiplas transacoes na mesma mensagem, inclui todas no array.

EXEMPLOS DE CATEGORIZACAO:
- Supermercado, restaurante, cafe, almoco, jantar, comida -> Alimentacao
- Uber, taxi, gasolina, combustivel, metro, autocarro, estacionamento -> Transportes
- Renda, agua, luz, gas, internet, condominio -> Habitacao
- Farmacia, medicamento, medico, consulta, dentista, ginasio -> Saude
- Cinema, Netflix, Spotify, jogos, bar, concerto, livro -> Entretenimento
- Salario, ordenado, freelance, rendimento -> Salario
- Roupa, sapatos, eletronicos, telemovel -> usa a categoria mais adequada das existentes

2. PERGUNTA FINANCEIRA — se pergunta sobre dados (quanto gastei, qual categoria, totais, etc.):
{{"intent":"question","query_type":"gastos_categoria|total_periodo|top_despesas|por_categoria","params":{{"category":"nome","period_start":"YYYY-MM-DD","period_end":"YYYY-MM-DD","limit":10}}}}
- Preenche apenas os params relevantes para a pergunta. Omite os que nao se aplicam.

3. CONSELHO/ANALISE — se pede dicas, sugestoes, analise de habitos:
{{"intent":"advice"}}

4. CONVERSA CASUAL — saudacoes, agradecimentos, conversa geral:
{{"intent":"chat","response":"texto da resposta em HTML simples (bold com <b>, italico com <i>)"}}

REGRAS:
- Responde APENAS com o objecto JSON. Sem explicacoes, sem markdown, sem ```json```.
- Para transacoes: nao inventes categorias; usa as existentes. Escolhe a categoria MAIS ESPECIFICA possivel.
- Se a mensagem e ambigua entre transacao e pergunta, prioriza transacao se contem um valor claro.
"""


async def ai_route_message(
    text: str,
    chat_id: str,
    user: models.User,
    workspace: models.Workspace,
    db: Session,
    t,
) -> dict:
    """
    AI Router: envia a mensagem ao GPT-4o-mini para classificar intent e extrair dados.
    Retorna dict com 'intent' e dados relevantes.
    Fallback: retorna intent='fallback' se GPT falhar (para usar parse_transaction).
    """
    if not settings.OPENAI_API_KEY:
        return {"intent": "fallback"}

    language = user.language if user.language else "pt"
    context = _build_financial_context(user, workspace, db)
    system_prompt = _build_system_prompt(context, language)

    try:
        from openai import OpenAI
        client = OpenAI(api_key=settings.OPENAI_API_KEY)

        # Incluir memória de conversa para contexto
        messages = [{"role": "system", "content": system_prompt}]
        memory = _get_memory(chat_id)
        if memory:
            messages.extend(memory)
        messages.append({"role": "user", "content": text})

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=messages,
            max_tokens=600,
            temperature=0.1,
        )

        raw = (response.choices[0].message.content or "").strip()
        logger.info("[AI Router] GPT raw response: %s", raw[:300])

        # Limpar resposta: por vezes GPT mete ```json ... ```
        if raw.startswith("```"):
            raw = re.sub(r'^```(?:json)?\s*', '', raw)
            raw = re.sub(r'\s*```$', '', raw)

        result = json.loads(raw)
        intent = result.get("intent", "chat")
        logger.info("[AI Router] intent=%s", intent)

        if intent == "transaction":
            transactions = result.get("transactions", [])
            if not transactions:
                return {"intent": "fallback"}
            return {"intent": "transaction", "transactions": transactions}

        elif intent == "question":
            query_type = result.get("query_type", "total_periodo")
            params = result.get("params", {})
            # Executar query na BD
            data_text = _query_financial_data(query_type, workspace, db, params)
            if not data_text or "nao encontrad" in data_text.lower():
                return {"intent": "chat", "response": t('ai_no_data')}
            # Segundo pedido ao GPT: formatar resposta bonita com os dados reais
            format_prompt = (
                f"Com base nestes dados financeiros reais do utilizador:\n\n{data_text}\n\n"
                f"Pergunta original: \"{text}\"\n\n"
                f"Formata uma resposta clara e bonita em HTML simples (usa <b> para bold, <i> para italico). "
                f"Inclui totais, medias se relevante, e comparacoes. Nao uses markdown, usa HTML. "
                f"Maximo 500 caracteres. {'Responde em portugues.' if language == 'pt' else 'Reply in English.'}"
            )
            fmt_response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": format_prompt}],
                max_tokens=300,
                temperature=0.3,
            )
            formatted = (fmt_response.choices[0].message.content or "").strip()
            # Limpar markdown residual
            formatted = re.sub(r'```(?:html)?\s*', '', formatted)
            formatted = re.sub(r'\s*```', '', formatted)
            return {"intent": "question", "response": t('ai_question_header') + formatted}

        elif intent == "advice":
            # Pedir conselho personalizado com contexto financeiro
            advice_prompt = (
                f"Es um consultor financeiro pessoal. Com base nestes dados:\n\n{context}\n\n"
                f"O utilizador pediu: \"{text}\"\n\n"
                f"Da conselhos praticos e personalizados. Identifica padroes nos gastos e sugere melhorias concretas. "
                f"Usa HTML simples (<b>, <i>). Maximo 600 caracteres. "
                f"Tom: calmo, zen, encorajador. {'Responde em portugues.' if language == 'pt' else 'Reply in English.'}"
            )
            adv_response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": advice_prompt}],
                max_tokens=350,
                temperature=0.5,
            )
            advice_text = (adv_response.choices[0].message.content or "").strip()
            advice_text = re.sub(r'```(?:html)?\s*', '', advice_text)
            advice_text = re.sub(r'\s*```', '', advice_text)
            return {"intent": "advice", "response": t('ai_advice_header') + advice_text}

        elif intent == "chat":
            chat_response = result.get("response", "")
            if not chat_response:
                chat_response = "🧘‍♂️"
            return {"intent": "chat", "response": chat_response}

        return {"intent": "fallback"}

    except json.JSONDecodeError as e:
        logger.warning("[AI Router] JSON decode error: %s | raw: %s", e, raw[:200] if 'raw' in dir() else '?')
        return {"intent": "fallback"}
    except ImportError:
        logger.warning("[AI Router] openai not installed")
        return {"intent": "fallback"}
    except Exception as e:
        err_str = str(e).lower()
        if "429" in err_str or "quota" in err_str or "rate" in err_str:
            logger.warning("[AI Router] OpenAI rate limit: %s", e)
            return {"intent": "gpt_rate_limited"}
        logger.error("[AI Router] Error: %s", e, exc_info=True)
        return {"intent": "fallback"}


def _handle_ai_transaction(
    transactions_data: list,
    chat_id: str,
    user: models.User,
    workspace: models.Workspace,
    db: Session,
    t,
    original_text: str = "",
) -> dict:
    """
    Processa transacoes extraidas pelo AI Router.
    Mapeia nomes de categoria para IDs reais e cria pendentes/transacoes.
    Retorna dict com status.
    """
    categories = db.query(models.Category).filter(
        models.Category.workspace_id == workspace.id
    ).all()

    # Fallback date from the original user message
    fallback_date = _parse_date_from_text(original_text) if original_text else None

    parsed_list = []
    for tx in transactions_data:
        amount = float(tx.get("amount", 0))
        if amount <= 0:
            continue
        description = (tx.get("description") or "Transacao").strip()[:255]
        tipo = tx.get("type", "expense")
        cat_name = (tx.get("category") or "").strip()

        # Parse date: GPT date field -> fallback from message text -> today
        tx_date = None
        gpt_date_str = tx.get("date")
        if gpt_date_str and gpt_date_str != "null":
            try:
                tx_date = datetime.strptime(str(gpt_date_str).strip(), "%Y-%m-%d").date()
            except (ValueError, TypeError):
                pass
        if not tx_date:
            tx_date = fallback_date or date.today()

        # Encontrar categoria por nome (exact -> fuzzy -> fallback)
        filtered = [c for c in categories if c.type == tipo]
        category_id = None
        for c in filtered:
            if c.name.lower() == cat_name.lower():
                category_id = c.id
                break
        if not category_id:
            match = find_best_category_match(cat_name, filtered, threshold=0.5)
            if match:
                category_id = match.id
        if not category_id and filtered:
            # Prefer specific categories over generic ones (e.g., "Despesas gerais")
            _generic = {'despesas gerais', 'general expenses', 'dépenses générales', 'outros', 'other'}
            specific = [c for c in filtered if c.name.lower() not in _generic]
            category_id = (specific[0] if specific else filtered[0]).id

        if not category_id:
            continue

        amount_cents = int(amount * 100)
        if tipo == "expense":
            amount_cents = -abs(amount_cents)
        else:
            amount_cents = abs(amount_cents)

        parsed_list.append({
            "amount": amount,
            "amount_cents": amount_cents,
            "description": description,
            "type": tipo,
            "category_id": category_id,
            "inference_source": "gpt4o_mini",
            "decision_reason": "ai_router",
            "needs_review": False,
            "transaction_date": tx_date,
        })

    if not parsed_list:
        return {"status": "no_valid_transactions"}

    # Se multiplas, usar batch flow
    if len(parsed_list) > 1:
        return {
            "status": "multiple",
            "parsed": {"multiple": True, "transactions": parsed_list},
        }

    # Transacao unica
    tx = parsed_list[0]
    cat = db.query(models.Category).filter(models.Category.id == tx["category_id"]).first()
    cat_name = cat.name if cat else "Outros"

    return {
        "status": "single",
        "parsed": {
            "amount": tx["amount"],
            "description": tx["description"],
            "type": tx["type"],
            "category_id": tx["category_id"],
            "inference_source": tx["inference_source"],
            "decision_reason": tx["decision_reason"],
            "needs_review": tx["needs_review"],
            "transaction_date": tx["transaction_date"],
        },
        "category_name": cat_name,
    }


# ==================== END AI ROUTER ====================


def normalize_text(text: str) -> str:
    """Normaliza texto removendo acentos e símbolos"""
    # Remove acentos
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    # Remove símbolos e converte para minúsculas
    text = re.sub(r'[^\w\s]', '', text.lower())
    return text

def similarity_score(str1: str, str2: str) -> float:
    """Calcula similaridade entre duas strings (0.0 a 1.0)"""
    return SequenceMatcher(None, str1, str2).ratio()


def _get_historical_canonical_forms(workspace_id: int, db: Session, limit: int = 400) -> List[tuple]:
    """
    Devolve formas canónicas (normalized, original) a partir das descrições históricas do workspace.
    Inclui descrições completas e palavras individuais (>=3 chars) para corrigir transcrições de voz.
    """
    # Descrições únicas ordenadas por frequência (mais usadas primeiro)
    rows = (
        db.query(models.Transaction.description, func.count(models.Transaction.id))
        .filter(
            models.Transaction.workspace_id == workspace_id,
            models.Transaction.description.isnot(None),
            models.Transaction.description != "",
        )
        .group_by(models.Transaction.description)
        .order_by(func.count(models.Transaction.id).desc())
        .limit(limit)
        .all()
    )
    seen_normalized = set()
    out = []
    for desc, _ in rows:
        if not desc or not desc.strip():
            continue
        d = desc.strip()[:255]
        norm = normalize_text(d)
        if norm not in seen_normalized:
            seen_normalized.add(norm)
            out.append((norm, d))
        for word in re.split(r"\s+", d):
            if len(word) >= 3 and re.search(r"[a-zA-Zàáâãäåèéêëìíîïòóôõöùúûüç]", word, re.IGNORECASE):
                w_norm = normalize_text(word)
                if w_norm not in seen_normalized:
                    seen_normalized.add(w_norm)
                    out.append((w_norm, word))
    return out


def correct_transcription_with_history(text: str, workspace_id: int, db: Session, threshold: float = 0.82) -> str:
    """
    Corrige o texto (ex.: transcrição de voz) usando descrições históricas do workspace.
    Substitui palavras semelhantes pela forma guardada no histórico (ortografia e maiúsculas corretas).
    """
    if not text or not text.strip():
        return text
    canonical = _get_historical_canonical_forms(workspace_id, db)
    if not canonical:
        return text
    tokens = re.split(r"\s+", text.strip())
    replacements = {}
    for token in tokens:
        if len(token) < 3 or re.match(r"^-?\d+([.,]\d+)?$", token) or token.lower() in ("euro", "euros", "eur", "€", "e"):
            continue
        if not re.search(r"[a-zA-Zàáâãäåèéêëìíîïòóôõöùúûüç]", token, re.IGNORECASE):
            continue
        t_norm = normalize_text(token)
        best_score = 0.0
        best_original = None
        for norm_form, original_form in canonical:
            score = similarity_score(t_norm, norm_form)
            if score >= threshold and score > best_score:
                best_score = score
                best_original = original_form
        if best_original is not None:
            replacements[token] = best_original
    if not replacements:
        return text
    # Aplicar substituições por palavra inteira (evitar alterar dentro de palavras)
    result = text
    for token, original in replacements.items():
        result = re.sub(rf"\b{re.escape(token)}\b", original, result, flags=re.IGNORECASE)
    if result != text:
        logger.info("[Histórico] Texto corrigido: '%s' -> '%s'", text[:80], result[:80])
    return result


def get_category_by_keyword(description: str, workspace_id, tipo: str, db: Session):
    """Se alguma palavra da descrição (normalizada) estiver em category_keywords, devolve category_id; senão None."""
    if not description or not hasattr(models, 'CategoryKeyword'):
        return None
    try:
        for word in description.split():
            if len(word) < 2:
                continue
            w_norm = normalize_text(word)
            if len(w_norm) < 2:
                continue
            kw = db.query(models.CategoryKeyword).filter(
                models.CategoryKeyword.workspace_id == workspace_id,
                models.CategoryKeyword.keyword == w_norm,
            ).first()
            if kw:
                cat = db.query(models.Category).filter(
                    models.Category.id == kw.category_id,
                    models.Category.type == tipo,
                ).first()
                if cat:
                    return cat.id
    except Exception as e:
        logger.warning("[Telegram] get_category_by_keyword falhou: %s", e)
    return None


def find_best_category_match(user_input: str, categories: List[models.Category], threshold: float = 0.6) -> Optional[models.Category]:
    """
    Encontra a categoria mais similar ao input do utilizador usando similaridade de strings.
    Retorna a categoria se a similaridade for >= threshold, caso contrário None.
    """
    user_input_normalized = normalize_text(user_input)
    best_match = None
    best_score = 0.0
    
    for cat in categories:
        cat_name_normalized = normalize_text(cat.name)
        
        # Calcular similaridade
        score = similarity_score(user_input_normalized, cat_name_normalized)
        
        # Também verificar se uma está contida na outra (match parcial)
        if user_input_normalized in cat_name_normalized or cat_name_normalized in user_input_normalized:
            score = max(score, 0.8)  # Boost para matches parciais
        
        # Verificar palavras individuais (útil para "aliments" vs "alimentacao")
        user_words = set(user_input_normalized.split())
        cat_words = set(cat_name_normalized.split())
        if user_words and cat_words:
            # Se há palavras em comum, aumentar score
            common_words = user_words.intersection(cat_words)
            if common_words:
                word_score = len(common_words) / max(len(user_words), len(cat_words))
                score = max(score, word_score * 0.9)
        
        # Verificar prefixo comum (útil para "aliments" vs "alimentacao")
        min_len = min(len(user_input_normalized), len(cat_name_normalized), 7)
        if min_len >= 4:
            if user_input_normalized[:min_len] == cat_name_normalized[:min_len]:
                score = max(score, 0.75)  # Boost para prefixos comuns
        
        if score > best_score:
            best_score = score
            best_match = cat
    
    # Só retornar se a similaridade for suficientemente alta
    if best_score >= threshold:
        logger.info(f"✓ Categoria encontrada por similaridade: '{best_match.name}' (score: {best_score:.2f}) para '{user_input}'")
        return best_match
    
    return None

def find_similar_transaction(text: str, workspace_id: uuid.UUID, db: Session, tipo: str) -> Optional[uuid.UUID]:
    """
    Busca transações similares no histórico (dados históricos).
    Prioridade máxima antes de IA: quanto mais hits, menos chamadas à OpenAI.
    Usa chave canonical (igual ao motor) para consistência.
    NÃO usa transações de seed (1 cêntimo).
    """
    cache_key = _description_cache_key(text)
    if not cache_key:
        return None
    words = set(cache_key.split())
    if not words:
        return None

    cutoff_date = date.today() - timedelta(days=180)
    transactions = db.query(models.Transaction).filter(
        models.Transaction.workspace_id == workspace_id,
        models.Transaction.transaction_date >= cutoff_date,
        models.Transaction.category_id.isnot(None),
    ).order_by(models.Transaction.transaction_date.desc()).limit(500).all()

    if tipo == "expense":
        transactions = [t for t in transactions if t.amount_cents < 0 and abs(t.amount_cents) != 1]
    else:
        transactions = [t for t in transactions if t.amount_cents > 0 and abs(t.amount_cents) != 1]

    best_match = None
    best_score = 0
    best_description = None

    for trans in transactions:
        if not trans.description:
            continue
        desc_can = _description_cache_key(trans.description)
        desc_words = set(desc_can.split())
        common = words.intersection(desc_words)
        if not common:
            continue
        score = len(common) + sum(2 for w in common if len(w) > 4)
        days_ago = (date.today() - trans.transaction_date).days
        score += 3 if days_ago <= 7 else (2 if days_ago <= 30 else (1 if days_ago <= 90 else 0))
        has_important = any(len(w) > 4 for w in common)
        # Relaxado: 1 palavra importante OU 2+ palavras em comum (favorece histórico antes de IA)
        accept = (has_important and len(common) >= 1) or len(common) >= 2
        if accept and score >= 2 and score > best_score:
            best_score = score
            best_match = trans.category_id
            best_description = trans.description

    if best_match:
        logger.info("Histórico similar: '%s' -> category_id (score=%s)", best_description, best_score)
    return best_match

def validate_email(email: str) -> bool:
    """Valida formato de email"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))


def _parse_date_from_text(text: str) -> Optional[date]:
    """
    Extrai uma data da mensagem para usar como transaction_date.
    Suporta: "ontem", "anteontem", "hoje", "yesterday", "today",
    "28/01", "28/01/25", "28/01/2025", "28-01", "dia 28", "28 jan", "28 janeiro".
    Retorna None se não encontrar data válida.
    """
    if not text or not text.strip():
        return None
    text_clean = text.strip()
    text_lower = text_clean.lower()
    today = date.today()

    # Relative dates: ontem, anteontem, hoje, yesterday, today
    if re.search(r'\b(?:ante[\s-]?ontem|antes\s+de\s+ontem|day\s+before\s+yesterday)\b', text_lower):
        return today - timedelta(days=2)
    if re.search(r'\bontem\b|\byesterday\b', text_lower):
        return today - timedelta(days=1)
    if re.search(r'\bhoje\b|\btoday\b', text_lower):
        return today

    # DD/MM ou DD/MM/YY ou DD/MM/YYYY
    m = re.search(r'\b(\d{1,2})[/\-](\d{1,2})(?:[/\-](\d{2,4}))?\b', text_clean)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), m.group(3)
        if y is None:
            year = today.year
        else:
            year = int(y)
            if year < 100:
                year += 2000 if year < 50 else 1900
        try:
            return date(year, mo, d)
        except ValueError:
            pass
    # "dia 28" / "day 28" (dia do mês atual)
    m = re.search(r'\b(?:dia|day)\s+(\d{1,2})\b', text_clean, re.IGNORECASE)
    if m:
        try:
            d = int(m.group(1))
            return date(today.year, today.month, d)
        except ValueError:
            pass
    # "28 jan" / "28 janeiro" / "15 dez"
    months_map = [
        (r'janeiro|january|jan', 1), (r'fevereiro|february|fev|feb', 2), (r'mar[cç]o|march|mar', 3),
        (r'abril|april|abr|apr', 4), (r'maio|may|mai', 5), (r'junho|june|jun', 6),
        (r'julho|july|jul', 7), (r'agosto|august|ago|aug', 8), (r'setembro|september|set|sep', 9),
        (r'outubro|october|out|oct', 10), (r'novembro|november|nov', 11), (r'dezembro|december|dez|dec', 12),
    ]
    m = re.search(r'\b(\d{1,2})\s+([a-zàáâãäåèéêëìíîïòóôõöùúûüç]+)\b', text_clean, re.IGNORECASE)
    if m:
        d = int(m.group(1))
        month_name = m.group(2).lower()
        for pattern, mo in months_map:
            if re.match(pattern, month_name, re.IGNORECASE):
                try:
                    return date(today.year, mo, d)
                except ValueError:
                    pass
                break
    return None

def _strip_date_from_description(description: str) -> str:
    """Remove padrões de data da descrição (ontem, anteontem, hoje, 28/01, dia 28, day 28, 28 jan, etc.) para não guardar na BD."""
    if not description:
        return description
    s = description.strip()
    s = re.sub(r'\b(?:ante[\s-]?ontem|antes\s+de\s+ontem|day\s+before\s+yesterday)\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\b(?:ontem|yesterday|hoje|today)\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\b\d{1,2}[/\-]\d{1,2}(?:[/\-]\d{2,4})?\b', '', s)
    s = re.sub(r'\b(?:dia|day)\s+\d{1,2}\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\b\d{1,2}\s+(?:jan|janeiro|fev|fevereiro|mar|março|marco|abr|abril|mai|maio|jun|junho|jul|julho|ago|agosto|set|setembro|out|outubro|nov|novembro|dez|dezembro|january|february|march|april|may|june|july|august|september|october|november|december)\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', ' ', s).strip()
    return s or description.strip()


# --- Números por extenso para voz (PT/EN) ---
# Unidades e 10–19 (ordem: mais longos primeiro para não partir "dezassete" em "dez"+"assete")
_NUMBER_WORDS_BASE = [
    ("dezassete", "17"), ("dezasseis", "16"), ("dezanove", "19"), ("dezoito", "18"),
    ("catorze", "14"), ("quatorze", "14"), ("quinze", "15"), ("treze", "13"),
    ("doze", "12"), ("onze", "11"), ("nove", "9"), ("oito", "8"), ("sete", "7"),
    ("seis", "6"), ("cinco", "5"), ("quatro", "4"), ("três", "3"), ("tres", "3"),
    ("dois", "2"), ("duas", "2"), ("uma", "1"), ("um", "1"),
    ("vinte", "20"), ("trinta", "30"), ("quarenta", "40"), ("cinquenta", "50"),
    ("sessenta", "60"), ("setenta", "70"), ("oitenta", "80"), ("noventa", "90"),
    ("cem", "100"), ("cento", "100"), ("duzentos", "200"), ("duzentas", "200"),
    ("trezentos", "300"), ("trezentas", "300"), ("quatrocentos", "400"), ("quatrocentas", "400"),
    ("quinhentos", "500"), ("quinhentas", "500"), ("seiscentos", "600"), ("seiscentas", "600"),
    ("setecentos", "700"), ("setecentas", "700"), ("oitocentos", "800"), ("oitocentas", "800"),
    ("novecentos", "900"), ("novecentas", "900"), ("mil", "1000"),
    ("seventeen", "17"), ("sixteen", "16"), ("nineteen", "19"), ("eighteen", "18"),
    ("fourteen", "14"), ("fifteen", "15"), ("thirteen", "13"), ("twelve", "12"),
    ("eleven", "11"), ("nine", "9"), ("eight", "8"), ("seven", "7"),
    ("six", "6"), ("five", "5"), ("four", "4"), ("three", "3"), ("two", "2"),
    ("one", "1"), ("twenty", "20"), ("thirty", "30"), ("forty", "40"), ("fifty", "50"),
    ("sixty", "60"), ("seventy", "70"), ("eighty", "80"), ("ninety", "90"),
    ("hundred", "100"), ("thousand", "1000"),
]
# Compostos PT: "vinte e um" .. "noventa e nove"
_PT_TENS = [("vinte", 20), ("trinta", 30), ("quarenta", 40), ("cinquenta", 50),
            ("sessenta", 60), ("setenta", 70), ("oitenta", 80), ("noventa", 90)]
_PT_ONES = [("um", 1), ("dois", 2), ("três", 3), ("tres", 3), ("quatro", 4), ("cinco", 5),
            ("seis", 6), ("sete", 7), ("oito", 8), ("nove", 9)]
_EN_TENS = [("twenty", 20), ("thirty", 30), ("forty", 40), ("fifty", 50),
            ("sixty", 60), ("seventy", 70), ("eighty", 80), ("ninety", 90)]
_EN_ONES = [("one", 1), ("two", 2), ("three", 3), ("four", 4), ("five", 5),
            ("six", 6), ("seven", 7), ("eight", 8), ("nine", 9)]


def _build_compound_patterns() -> List[tuple]:
    """Gera padrões compostos (ex.: 'vinte e cinco' -> 25) para PT e EN."""
    out = []
    for (t_name, t_val), (o_name, o_val) in [(t, o) for t in _PT_TENS for o in _PT_ONES]:
        out.append((f"{t_name} e {o_name}", str(t_val + o_val)))
    for (t_name, t_val), (o_name, o_val) in [(t, o) for t in _EN_TENS for o in _EN_ONES]:
        out.append((f"{t_name} {o_name}", str(t_val + o_val)))
    return out


_COMPOUND_PATTERNS = _build_compound_patterns()


def _normalize_number_words(text: str) -> str:
    """
    Substitui números por extenso (e compostos) por dígitos para melhor parse de voz.
    - Compostos: "vinte e cinco" -> 25, "thirty five" -> 35
    - Centenas: "cento e 25" -> 125, "mil e 500" -> 1500
    - "N e meio" -> N.5 (ex.: "quinze e meio" -> 15.5)
    - "dez" não é substituído após um dia (ex.: "28 dez" = dezembro)
    """
    if not text or not text.strip():
        return text
    s = text
    # 1) Compostos primeiro (vinte e cinco -> 25, thirty five -> 35)
    for phrase, digit in _COMPOUND_PATTERNS:
        s = re.sub(rf"\b{re.escape(phrase)}\b", digit, s, flags=re.IGNORECASE)
    # 2) Centena + e + número ANTES de substituir "cento"/"mil" (cento e 25 -> 125)
    for prefix, base in [
        ("cento e", 100), ("cem e", 100), ("duzentos e", 200), ("duzentas e", 200),
        ("trezentos e", 300), ("trezentas e", 300), ("quatrocentos e", 400), ("quatrocentas e", 400),
        ("quinhentos e", 500), ("quinhentas e", 500), ("seiscentos e", 600), ("seiscentas e", 600),
        ("setecentos e", 700), ("setecentas e", 700), ("oitocentos e", 800), ("oitocentas e", 800),
        ("novecentos e", 900), ("novecentas e", 900), ("mil e", 1000),
        ("one hundred and", 100), ("two hundred and", 200), ("three hundred and", 300),
    ]:
        def _repl(m, b=base):
            try:
                return str(b + int(m.group(1)))
            except ValueError:
                return m.group(0)
        s = re.sub(rf"\b{re.escape(prefix)}\s+(\d+)\b", lambda m, _b=base: _repl(m, _b), s, flags=re.IGNORECASE)
    # 3) Palavras simples (exceto "dez", tratado em 5)
    for word, digit in _NUMBER_WORDS_BASE:
        if word.lower() == "dez":
            continue
        s = re.sub(rf"\b{re.escape(word)}\b", digit, s, flags=re.IGNORECASE)
    # 4) "N e meio" -> N.5 (ex.: "quinze e meio" já é "15 e meio" após passo 3)
    s = re.sub(r"\b(\d+)\s+e\s+meio\b", r"\1.5", s, flags=re.IGNORECASE)
    # 5) "dez" -> 10, exceto em datas ("28 dez") e em "dezembro"
    s = re.sub(r"(?<!\s\d)\bdez\b(?!embro)", "10", s, flags=re.IGNORECASE)
    return s


def _remove_dates_for_value_parsing(text: str) -> str:
    """
    Remove padrões de data do texto antes de extrair valores monetários.
    Evita que "01/02/2026" seja interpretado como vários valores (1, 2, 202, 6).
    Substitui datas por espaço para manter posições coerentes.
    """
    if not text:
        return text
    s = text
    # DD/MM/YYYY, DD/MM/YY, DD/MM, DD-MM (substituir por espaço)
    s = re.sub(r'\b\d{1,2}[/\-]\d{1,2}(?:[/\-]\d{2,4})?\b', ' ', s)
    # "dia 28" / "day 28"
    s = re.sub(r'\b(?:dia|day)\s+\d{1,2}\b', ' ', s, flags=re.IGNORECASE)
    # "28 jan" / "28 janeiro" / etc.
    s = re.sub(r'\b\d{1,2}\s+(?:jan|janeiro|fev|fevereiro|mar|mar[cç]o|marco|abr|abril|mai|maio|jun|junho|jul|julho|ago|agosto|set|setembro|out|outubro|nov|novembro|dez|dezembro|january|february|march|april|may|june|july|august|september|october|november|december)\b', ' ', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def parse_transaction(
    text: str,
    workspace: models.Workspace,
    db: Session,
    default_category_id: Optional[uuid.UUID] = None,
) -> Optional[Dict]:
    """
    Extrai valor, tipo e categoria de uma mensagem de texto.
    Suporta múltiplas transações separadas por espaço.
    Aceita data na mensagem: "Almoço 15€ 28/01", "Almoço 28/01 15€", "dia 28 Almoço 15€".
    Normaliza números por extenso (ex.: quinze euros) para melhor suporte a voz.
    default_category_id: se definido (/categoria), usa esta categoria para todas as transações da mensagem.
    """
    # Normalizar números por extenso (voz: "quinze euros" -> "15 euros")
    text = _normalize_number_words(text)
    # Data opcional na mensagem (aplica-se a todas as transações da mensagem)
    parsed_date = _parse_date_from_text(text)
    default_transaction_date = parsed_date if parsed_date else date.today()

    # Remover datas do texto antes de extrair valores, para "Almoco 15€ 01/02/2026" não virar vários valores (1, 2, 202, 6)
    text_for_values = _remove_dates_for_value_parsing(text)

    # Suporta múltiplas transações: "Almoço 15€ Gasolina 10€"
    transactions = []
    
    # Regex para encontrar valores monetários (inclui -15€)
    # Só considerar valor se estiver associado a €/eur/euros (evita apanhar números soltos de datas)
    valor_pattern = r'(-?\d{1,3}(?:[.\s]\d{3})*(?:[.,]\d+)?)\s*(?:€|eur|euros?|e)\b'
    valor_matches = list(re.finditer(valor_pattern, text_for_values, re.IGNORECASE))
    if not valor_matches:
        # Fallback: sem símbolo € (ex.: "15" no fim de frase)
        valor_pattern = r'(-?\d{1,3}(?:[.\s]\d{3})*(?:[.,]\d+)?)\s*(?:€|eur|euros?|e)?'
        valor_matches = list(re.finditer(valor_pattern, text_for_values, re.IGNORECASE))
    if not valor_matches:
        return None
    
    # Identificar tipo (despesa ou receita)
    text_lower = text.lower()
    income_keywords = [
        'recebi', 'salário', 'ordenado', 'ganhei', 'vendi', 'rendimento',
        'bonus', 'vencimento', 'reembolso', 'subsídio', 'prémio', 'premio',
        'venda', 'cashback', 'entrada', 'received', 'salary', 'income', 'refund'
    ]
    expense_keywords = [
        'pago', 'paguei', 'debitado', 'saída', 'gastei', 'paid', 'spent', 'debit'
    ]
    vault_withdrawal_keywords = [
        'retirar', 'resgate', 'retirei', 'sacar', 'levantei', 'withdraw', 'withdrawal'
    ]
    if any(k in text_lower for k in income_keywords):
        tipo = "income"
    elif any(k in text_lower for k in expense_keywords):
        tipo = "expense"
    else:
        tipo = "expense"
    is_vault_withdrawal = any(k in text_lower for k in vault_withdrawal_keywords)
    
    # Buscar categorias do workspace
    categories = db.query(models.Category).filter(
        models.Category.workspace_id == workspace.id,
        models.Category.type == tipo
    ).all()
    
    if not categories:
        return None
    
    # Verificar se o utilizador especificou uma categoria na mensagem
    # Formato: "Bolachas - Alimentação 100€" ou "Bolachas - alimentos 100€"
    text_lower_normalized = normalize_text(text)
    specified_category = None
    specified_category_name = None
    
    # Primeiro, verificar se há um hífen separando descrição da categoria
    # Formato: "Descrição - Categoria Valor€"
    if ' - ' in text or ' -' in text or '- ' in text:
        # Dividir por hífen
        parts = re.split(r'\s*-\s*', text, 1)
        if len(parts) == 2:
            # parts[0] = descrição, parts[1] = categoria + valor
            category_part = parts[1]
            # Remover o valor monetário da parte da categoria
            category_part_clean = re.sub(r'\s*\d+[.,\s]*\d*\s*(?:€|eur|euros|e)?', '', category_part, flags=re.IGNORECASE).strip()
            category_part_normalized = normalize_text(category_part_clean)
            
            # Usar similaridade de strings para encontrar a melhor correspondência
            specified_category = find_best_category_match(category_part_clean, categories, threshold=0.6)
            if specified_category:
                specified_category_name = specified_category.name
    
    # Se não encontrou com hífen, verificar match direto no texto completo usando similaridade
    if not specified_category:
        # Primeiro, verificar match exato (mais rápido)
        for cat in categories:
            cat_name_normalized = normalize_text(cat.name)
            if cat_name_normalized in text_lower_normalized:
                specified_category = cat
                specified_category_name = cat.name
                logger.info(f"✓ Categoria especificada na mensagem (match direto): '{cat.name}' (id: {cat.id})")
                break
        
        # Se não encontrou match exato, usar similaridade em palavras do texto
        if not specified_category:
            text_words = text_lower_normalized.split()
            for word in text_words:
                if len(word) >= 4:  # Só verificar palavras com pelo menos 4 caracteres
                    match = find_best_category_match(word, categories, threshold=0.7)
                    if match:
                        specified_category = match
                        specified_category_name = match.name
                        logger.info(f"✓ Categoria encontrada por similaridade na palavra '{word}': '{match.name}'")
                        break
    
    # Limites de valor (evitar input acidental)
    MAX_AMOUNT = 999_999.99

    # Processar cada valor encontrado
    for i, valor_match in enumerate(valor_matches):
        # Extrair valor (suporta -15€)
        valor_str = valor_match.group(1).replace(' ', '').replace('.', '').replace(',', '.')
        try:
            amount = float(valor_str)
        except ValueError:
            continue
        if abs(amount) > MAX_AMOUNT:
            continue
        # Valor negativo explícito = despesa
        if amount < 0:
            amount = abs(amount)
            tipo = "expense"
        
        # Extrair descrição (texto antes do valor, ou texto entre valores) — usar text_for_values (posições dos matches)
        if ' - ' in text_for_values or ' -' in text_for_values or '- ' in text_for_values:
            text_parts = re.split(r'\s*-\s*', text_for_values, 1)
            if len(text_parts) == 2:
                first_part = text_parts[0].strip()
                description = re.sub(r'\s*\d+[.,\s]*\d*\s*(?:€|eur|euros|e)?', '', first_part, flags=re.IGNORECASE).strip()
                logger.info(f"Descrição após separar por hífen: '{description}'")
            else:
                start_pos = valor_matches[i-1].end() if i > 0 else 0
                end_pos = valor_match.start()
                description = text_for_values[start_pos:end_pos].strip()
        else:
            start_pos = valor_matches[i-1].end() if i > 0 else 0
            end_pos = valor_match.start()
            description = text_for_values[start_pos:end_pos].strip()
        
        # Limpar separadores de voz em múltiplas transações: " e gasolina", ", gasolina", " e "
        description = re.sub(r"^[\s,]+", "", description)
        description = re.sub(r"[\s,]+$", "", description)
        description = re.sub(r"^\s*e\s+", "", description, flags=re.IGNORECASE)
        description = re.sub(r"\s+e\s*$", "", description, flags=re.IGNORECASE)
        description = description.strip()
        
        # Limpar descrição (remover categoria se foi especificada sem hífen)
        words_to_remove = ['€', 'euro', 'euros', 'eur', 'e', 'gastei', 'paguei', 'recebi', 
                          'em', 'no', 'na', 'de', 'do', 'da', 'com', 'para']
        
        # Se categoria foi especificada (sem hífen), removê-la da descrição (incluindo variações parciais)
        if specified_category and not (' - ' in text_for_values or ' -' in text_for_values or '- ' in text_for_values):
                desc_words = description.split()
                category_name_normalized = normalize_text(specified_category.name)
                # Remover palavras que correspondem à categoria (exato ou parcial)
                filtered_words = []
                for word in desc_words:
                    word_normalized = normalize_text(word)
                    # Verificar se a palavra é parte da categoria ou vice-versa
                    is_category_word = (
                        word_normalized == category_name_normalized or
                        category_name_normalized in word_normalized or
                        word_normalized in category_name_normalized
                    )
                    if not is_category_word:
                        filtered_words.append(word)
                description = " ".join(filtered_words).strip()
                logger.info(f"Descrição após remover categoria '{specified_category.name}': '{description}'")
        
        desc_words = description.split()
        final_desc_words = [w for w in desc_words if w.lower() not in words_to_remove]
        
        if final_desc_words:
            description = " ".join(final_desc_words).strip()
        else:
            description = "Transação Telegram"
        description = _strip_date_from_description(description)[:255]
        if not description:
            description = "Transação Telegram"
        
        inference_source = "fallback"
        needs_review = True
        decision_reason = ""
        category_id = None
        suggested_category_name = None
        # Categoria por defeito (/categoria ou /definir): usar se válida para este tipo
        if default_category_id:
            def_cat = next((c for c in categories if c.id == default_category_id), None)
            if def_cat:
                category_id = default_category_id
                inference_source = "telegram_default"
                needs_review = False
                decision_reason = "telegram_default"
        # Se categoria foi especificada na mensagem, usar diretamente (SEM ir ao motor)
        if specified_category:
            category_id = specified_category.id
            inference_source = "explicit"
            needs_review = False
            decision_reason = f"explicit:{specified_category_name}"
            logger.info(f"✓ Usando categoria especificada pelo utilizador: '{specified_category_name}' (id: {category_id})")
        elif not category_id:
            # Keywords por categoria (workspace): palavra → categoria
            keyword_cat_id = get_category_by_keyword(description, workspace.id, tipo, db)
            if keyword_cat_id:
                category_id = keyword_cat_id
                inference_source = "keyword"
                decision_reason = "keyword_telegram"
                needs_review = False
        if not category_id:
            # Prioridade: cache (exato) → histórico/similaridade → motor SEM IA → (opcional) IA só se nada acertar
            cache_key = _description_cache_key(description)
            inference_source = "legacy_fallback"
            decision_reason = "legacy_fallback"

            # 1) Cache privado/global (chave canonical = mesmo que o motor)
            if cache_key:
                category_id = get_cached_category(cache_key, workspace.id, tipo, categories, db)
                if category_id:
                    inference_source = "cache_private"
                    decision_reason = "cache_telegram"

            # 2) Similaridade com histórico (transações passadas do utilizador)
            if not category_id:
                category_id = find_similar_transaction(description, workspace.id, db, tipo)
                if category_id:
                    inference_source = "history_similarity"
                    decision_reason = "history_similarity_telegram"

            # 3) Motor de categorização SEM IA (regras, token-scoring, cache do motor, similaridade do motor)
            if not category_id:
                try:
                    from ..core.categorization_engine import infer_category
                    from ..core.config import settings
                    cat_id, source, needs_review, conf, reason, explain = infer_category(
                        description,
                        workspace.id,
                        tipo,
                        categories,
                        db,
                        models,
                        settings,
                        explicit_category_id=None,
                        use_gemini=False,  # IA só para imagens; texto não chama OpenAI
                    )
                    category_id = cat_id
                    inference_source = source
                    decision_reason = reason
                except Exception as e:
                    logger.warning("Motor de categorização falhou: %s", e)

            # 4) Último recurso: IA (só se não houver cache/histórico/motor) e circuit-breaker fechado
            suggested_category_name = None
            if not category_id:
                try:
                    from ..core.categorization_engine import check_gemini_circuit_breaker
                    if check_gemini_circuit_breaker(db, models, workspace.id):
                        logger.warning("Circuit-breaker IA aberto no Telegram; não chamar OpenAI")
                        _generic = {'despesas gerais', 'general expenses', 'dépenses générales', 'outros', 'other'}
                        _specific = [c for c in categories if c.name.lower() not in _generic]
                        category_id = (_specific[0] if _specific else categories[0]).id if categories else None
                        inference_source = "fallback"
                        decision_reason = "fallback:circuit_breaker"
                    else:
                        cat_id, suggested_name = categorize_with_ai(description, categories, tipo, text, workspace.id, db)
                        if cat_id:
                            category_id = cat_id
                            inference_source = "openai"
                            decision_reason = "openai:last_resort"
                            cache_key_save = _description_cache_key(description)
                            if cache_key_save:
                                cat_obj = next((c for c in categories if c.id == category_id), None)
                                save_cached_category(cache_key_save, workspace.id, category_id, cat_obj.name if cat_obj else "Outros", tipo, db, is_common=True)
                        elif suggested_name:
                            suggested_category_name = suggested_name[:100]
                            inference_source = "openai"
                            decision_reason = "openai:suggest_new"
                except Exception as e:
                    logger.warning("IA (último recurso) falhou: %s", e)

            # 5) Fallback final: prefer specific category over generic (e.g., "Despesas gerais")
            if not category_id and categories and not suggested_category_name:
                logger.info("Sem cache/histórico/motor/IA: usando fallback do tipo '%s'", tipo)
                _generic = {'despesas gerais', 'general expenses', 'dépenses générales', 'outros', 'other'}
                _specific = [c for c in categories if c.name.lower() not in _generic]
                category_id = (_specific[0] if _specific else categories[0]).id
                inference_source = "fallback"
                decision_reason = "fallback:first_category"
        
        # Verificar se a categoria é de vault (investimento/emergência)
        category_obj = db.query(models.Category).filter(models.Category.id == category_id).first() if category_id else None
        if not category_obj and category_id:
            category_obj = next((cat for cat in categories if cat.id == category_id), None)
        is_vault_category = category_obj and category_obj.vault_type != 'none' if category_obj else False
        
        transactions.append({
            "amount": amount,
            "description": description[:255],
            "type": tipo,
            "category_id": category_id,
            "inference_source": inference_source,
            "needs_review": needs_review,
            "decision_reason": decision_reason,
            "is_vault": is_vault_category,
            "is_vault_withdrawal": is_vault_withdrawal if is_vault_category else False,
            "suggested_category_name": suggested_category_name,
            "transaction_date": default_transaction_date,
        })
    
    # Retornar primeira transação ou lista se múltiplas
    if len(transactions) == 1:
        return transactions[0]
    return {"multiple": True, "transactions": transactions}


def _description_cache_key(description: str) -> str:
    """
    Chave de cache alinhada com o motor de categorização (canonicalize).
    Assim o cache do Telegram e o do motor são o mesmo → menos IA.
    """
    from ..core.categorization_engine import canonicalize
    return canonicalize(description) or ""


def get_cached_category(description_normalized: str, workspace_id: uuid.UUID, tipo: str, categories: List[models.Category], db: Session) -> Optional[uuid.UUID]:
    """
    Verifica se existe uma categorização em cache para esta descrição.
    description_normalized deve ser a chave canonical (canonicalize(description)) para alinhar com o motor.
    """
    # 1. Verificar cache privado do workspace
    cache_entry = db.query(models.CategoryMappingCache).filter(
        models.CategoryMappingCache.workspace_id == workspace_id,
        models.CategoryMappingCache.description_normalized == description_normalized,
        models.CategoryMappingCache.transaction_type == tipo
    ).first()
    
    if cache_entry and cache_entry.category_id:
        # Atualizar contador e última utilização
        cache_entry.usage_count += 1
        cache_entry.last_used_at = datetime.now(timezone.utc)
        db.commit()
        logger.info(f"Cache privado hit: '{description_normalized}' -> '{cache_entry.category_id}' (usado {cache_entry.usage_count}x)")
        return cache_entry.category_id
    
    # 2. Verificar cache global (partilhado entre utilizadores)
    global_cache = db.query(models.CategoryMappingCache).filter(
        models.CategoryMappingCache.is_global == True,
        models.CategoryMappingCache.workspace_id.is_(None),
        models.CategoryMappingCache.description_normalized == description_normalized,
        models.CategoryMappingCache.transaction_type == tipo
    ).first()
    
    if global_cache:
        # Procurar categoria com o mesmo nome no workspace atual
        category_name = global_cache.category_name
        for cat in categories:
            if cat.name == category_name and cat.type == tipo:
                # Atualizar contador do cache global
                global_cache.usage_count += 1
                global_cache.last_used_at = datetime.now(timezone.utc)
                db.commit()
                logger.info(f"Cache global hit: '{description_normalized}' -> '{category_name}' (usado {global_cache.usage_count}x globalmente)")
                return cat.id
    
    return None

def save_cached_category(description_normalized: str, workspace_id: uuid.UUID, category_id: uuid.UUID, category_name: str, tipo: str, db: Session, is_common: bool = False):
    """
    Guarda uma categorização no cache para reutilização futura.
    Se is_common=True, guarda também no cache global (partilhado).
    """
    try:
        # 1. Guardar no cache privado do workspace
        existing = db.query(models.CategoryMappingCache).filter(
            models.CategoryMappingCache.workspace_id == workspace_id,
            models.CategoryMappingCache.description_normalized == description_normalized,
            models.CategoryMappingCache.transaction_type == tipo
        ).first()
        
        if existing:
            # Atualizar existente
            existing.category_id = category_id
            existing.category_name = category_name
            existing.usage_count += 1
            existing.last_used_at = datetime.now(timezone.utc)
            if hasattr(existing, 'confidence') and existing.confidence is not None:
                existing.confidence = min(1.0, float(existing.confidence) + 0.05)
        else:
            # Criar novo
            cache_entry = models.CategoryMappingCache(
                workspace_id=workspace_id,
                description_normalized=description_normalized,
                category_id=category_id,
                category_name=category_name,
                transaction_type=tipo,
                is_global=False,
                confidence=0.9
            )
            db.add(cache_entry)
        
        # 2. Se for uma categoria comum (ex: "Alimentação", "Transportes"), guardar também no cache global
        # Categorias comuns que todos os utilizadores têm
        common_category_names = ['Alimentação', 'Transportes', 'Habitação', 'Saúde', 'Entretenimento', 'Salário']
        
        if is_common or category_name in common_category_names:
            global_existing = db.query(models.CategoryMappingCache).filter(
                models.CategoryMappingCache.is_global == True,
                models.CategoryMappingCache.workspace_id.is_(None),
                models.CategoryMappingCache.description_normalized == description_normalized,
                models.CategoryMappingCache.transaction_type == tipo
            ).first()
            
            if not global_existing:
                # Criar cache global (sem workspace_id, sem category_id específico)
                global_cache = models.CategoryMappingCache(
                    workspace_id=None,
                    description_normalized=description_normalized,
                    category_id=None,  # Não precisa de category_id específico (cada workspace tem o seu)
                    category_name=category_name,
                    transaction_type=tipo,
                    is_global=True,
                    confidence=0.95
                )
                db.add(global_cache)
                logger.info(f"Categoria comum guardada no cache global: '{description_normalized}' -> '{category_name}'")
        
        db.commit()
        logger.info(f"Categoria guardada no cache privado: '{description_normalized}' -> '{category_id}'")
    except Exception as e:
        logger.error(f"Erro ao guardar no cache: {str(e)}")
        db.rollback()

def categorize_with_ai(text: str, categories: List[models.Category], tipo: str, original_text: str, workspace_id: uuid.UUID, db: Session) -> tuple:
    """
    Usa OpenAI GPT-4o-mini para categorizar a transação quando não encontra no cache.
    Retorna (category_id, suggested_category_name).
    - Se a IA acertar numa categoria existente: (category_id, None).
    - Se a IA sugerir um nome que não existe: (None, ai_category_name) para o bot perguntar se quer criar.
    """
    if not settings.OPENAI_API_KEY:
        logger.warning("OPENAI_API_KEY não configurada. Não é possível usar IA para categorizar.")
        return (None, None)
    
    filtered_categories = [cat for cat in categories if cat.type == tipo]
    if not filtered_categories:
        logger.warning(f"Nenhuma categoria do tipo '{tipo}' disponível")
        return (None, None)
    
    try:
        from openai import OpenAI
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
        
        categories_list = [cat.name for cat in filtered_categories]
        categories_text = ", ".join(categories_list)
        
        prompt = f"""Categoriza: "{original_text}"

Categorias: {categories_text}

Responde APENAS com o nome exato da categoria:"""
        
        logger.info(f"Consultando OpenAI: '{original_text}' -> {categories_list}")
        
        try:
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=20,
                temperature=0.1,
            )
            ai_category_name = ""
            if response.choices and response.choices[0].message.content:
                ai_category_name = response.choices[0].message.content.strip()
            logger.info(f"Resposta OpenAI: '{ai_category_name}'")
            
            if not ai_category_name:
                return (None, None)
            
            # 1) Match exato
            for cat in filtered_categories:
                if cat.name.lower() == ai_category_name.lower():
                    logger.info(f"Match exato: '{cat.name}' (id: {cat.id})")
                    return (cat.id, None)
            
            # 2) Match fuzzy (find_best_category_match)
            fuzzy_match = find_best_category_match(ai_category_name, filtered_categories, threshold=0.6)
            if fuzzy_match:
                logger.info(f"Match fuzzy: '{fuzzy_match.name}' para IA '{ai_category_name}'")
                return (fuzzy_match.id, None)
            
            # 3) Match parcial (substring)
            for cat in filtered_categories:
                if cat.name.lower() in ai_category_name.lower() or ai_category_name.lower() in cat.name.lower():
                    logger.info(f"Match parcial: '{cat.name}' (id: {cat.id})")
                    return (cat.id, None)
            
            # 4) Primeira palavra
            first_word = ai_category_name.split()[0]
            for cat in filtered_categories:
                if first_word.lower() in cat.name.lower():
                    logger.info(f"Match por palavra: '{cat.name}' (id: {cat.id})")
                    return (cat.id, None)
            
            # Nenhum match: sugerir criar a categoria com o nome da IA
            logger.warning(f"Nenhuma categoria encontrada para: '{ai_category_name}' -> sugerir criar")
            return (None, ai_category_name[:100])
                        
        except Exception as e:
            err_str = (str(e) or "").lower()
            if (
                "429" in err_str or "quota" in err_str or "rate_limit" in err_str
                or "rate limit" in err_str or "exceeded" in err_str
            ):
                logger.warning(f"OpenAI indisponível (quota/limite): {str(e)}")
                raise AIUnavailableError(str(e)) from e
            logger.error(f"Erro ao usar OpenAI: {str(e)}")
            return (None, None)
        
    except ImportError:
        logger.warning("openai não instalado. Instale com: pip install openai")
        return (None, None)
    except AIUnavailableError:
        raise
    except Exception as e:
        logger.error(f"Erro na categorização IA: {str(e)}")
        return (None, None)


def parse_document_with_openai(
    extracted_text: str,
    workspace: models.Workspace,
    db: Session,
    default_category_id: Optional[uuid.UUID],
) -> Optional[Dict]:
    """
    Usa OpenAI para extrair transações de texto longo (PDF/CSV/Excel).
    Devolve {"transactions": [{"amount", "description", "type", "date", "category"}]} compatível com _build_parsed_from_photo_result,
    ou None se falhar ou OPENAI_API_KEY não estiver definido.
    """
    if not getattr(settings, "OPENAI_API_KEY", None) or not (settings.OPENAI_API_KEY or "").strip():
        return None
    text = (extracted_text or "").strip()
    if len(text) < 100:
        return None
    categories = db.query(models.Category).filter(models.Category.workspace_id == workspace.id).all()
    if not categories:
        return None
    expense_names = [c.name for c in categories if getattr(c, "type", "expense") == "expense"]
    income_names = [c.name for c in categories if getattr(c, "type", "income") == "income"]
    cats_expense = ", ".join(expense_names[:40]) if expense_names else "(nenhuma)"
    cats_income = ", ".join(income_names[:40]) if income_names else "(nenhuma)"
    default_cat_name = None
    if default_category_id:
        default_cat = next((c for c in categories if c.id == default_category_id), None)
        if default_cat:
            default_cat_name = default_cat.name
    max_chars = 11000
    if len(text) > max_chars:
        text = text[:max_chars] + "\n[... texto truncado ...]"
    today_str = date.today().strftime("%Y-%m-%d")
    # Categoria padrão para transferências recebidas (receita) e enviadas (despesa)
    first_income_cat = income_names[0] if income_names else "Outros"
    first_expense_cat = expense_names[0] if expense_names else "Outros"
    _aliment = re.compile(r"alimenta[cç]", re.IGNORECASE)
    # Nomes da categoria "Despesas gerais" por língua (criada por defeito) — preferir para transferências "Para"
    GENERAL_EXPENSE_NAMES = {"Despesas gerais", "General expenses", "Dépenses générales"}
    transfer_expense_cat = next(
        (n for n in expense_names if n and n.strip() and n in GENERAL_EXPENSE_NAMES),
        None,
    )
    if not transfer_expense_cat:
        transfer_expense_cat = next(
            (n for n in expense_names if ("transfer" in n.lower() or "transferência" in n.lower()) and not _aliment.search(n)),
            None,
        )
    if not transfer_expense_cat:
        transfer_expense_cat = next(
            (n for n in expense_names if _aliment.search(n) is None and (n and n.strip())),
            first_expense_cat,
        )
    if not transfer_expense_cat or _aliment.search(transfer_expense_cat or ""):
        transfer_expense_cat = first_expense_cat

    prompt = f"""Analisa o seguinte texto extraído de um ficheiro (extrato bancário, lista de movimentos, CSV ou PDF) e extrai TODAS as transações, sem omitir nenhuma.

Regras OBRIGATÓRIAS:
- TRANSFERÊNCIAS RECEBIDAS (dinheiro que entrou) → type "income". Se a descrição não indicar destinatário (ex. só "Trf. Mb Way De"), usa category "{first_income_cat}" ou Salário. Se indicar origem (ex. "De Empresa X"), podes usar a categoria mais adequada das Receitas.
- TRANSFERÊNCIAS ENVIADAS (dinheiro que saiu) → type "expense". Se a descrição for genérica (ex. só "Trf. Mb Way Para" sem nome) usa category "{transfer_expense_cat}". Se a descrição indicar destinatário/comerciante (ex. "Trf. Mb Way Para McDonald's", "Para Continente", "Transfer to Restaurant"), usa a categoria adequada às Despesas: Alimentação para restaurantes/supermercados, Transportes para combustível/uber, etc. NUNCA uses Alimentação para transferências genéricas sem destinatário.
- Compras, Via Verde, Google One, etc. = expense com categoria adequada (Transportes, Alimentação, etc.).
- Se uma linha tiver dois valores em €, o primeiro é o valor do movimento.
- amount: número positivo (valor absoluto em euros).
- date: OBRIGATÓRIO extrair a data do documento quando visível. Formato "DD mmm YYYY" (ex: 06 fev 2026) → converte para YYYY-MM-DD (2026-02-06). Se a data aparecer na linha ou na secção, usa-a. Só usa "{today_str}" se não houver data no texto.
- description: descrição curta; mantém "Trf. Mb Way De Nome", "Via Verde", etc.

Categorias (nome EXATO):
Despesas: {cats_expense}
Receitas: {cats_income}

Responde APENAS com JSON válido, sem markdown:
{{"transactions":[{{"amount":n,"description":"...","type":"expense|income","date":"YYYY-MM-DD","category":"NomeExato"}}]}}

Texto do ficheiro:
---
{text}
---"""

    try:
        from openai import OpenAI
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=4000,
            temperature=0.1,
        )
        raw = (response.choices[0].message.content or "").strip() if response.choices else ""
        if not raw:
            return None
        clean = re.search(r"\{[\s\S]*\}", raw)
        if clean:
            raw = clean.group(0)
        data = json.loads(raw)
        items = data.get("transactions")
        if not isinstance(items, list) or len(items) == 0:
            return None
        out = []
        # Padrões transferência RECEBIDA (receita) — varia por banco/país (PT, ES, EN, FR, etc.)
        trf_recebida_patterns = [
            r"\b(?:trf\.?|mb\s*way|trf\.?\s*imed\.?)\s+de\b",           # PT: Trf. Mb Way De, Trf.imed. De
            r"\btransfer(?:ência|encia)?\s+de\b",                      # PT/ES: Transferência de, Transferencia de
            r"\brecebido\s+de\b",                                       # PT: Recebido de
            r"\b(?:recibido|recibo)\s+de\b",                            # ES: Recibido de
            r"\bbizum\s+de\b",                                          # ES: Bizum de (recebimento)
            r"\bpix\s+(?:recebido|entrada|de)\b",                       # BR: Pix recebido / de
            r"\btransfer\s+from\b",                                     # EN: Transfer from
            r"\breceived\s+from\b",                                     # EN: Received from
            r"\bincoming\s+transfer\b",                                  # EN: Incoming transfer
            r"\bcredit\s+from\b",                                       # EN: Credit from
            r"\bvirement\s+(?:reçu|de)\b",                              # FR: Virement reçu/de
            r"\btransfert\s+de\b",                                     # FR: Transfert de
            r"\breçu\s+de\b",                                           # FR: Reçu de
            r"\b(?:mb\s*way|mbway)\s+recebido\b",                       # PT: MB Way recebido
            r"\bentrada\s+de\s+transferência\b",                        # PT: Entrada de transferência
            r"\b(?:depósito|deposit)\s+de\b",                           # PT/EN: Depósito de
            r"\bcredito\s+(?:de|from)\b",                               # PT/ES: Crédito de
        ]
        trf_recebida = re.compile("|".join(f"(?:{p})" for p in trf_recebida_patterns), re.IGNORECASE)
        # Padrões transferência ENVIADA (despesa)
        trf_enviada_patterns = [
            r"\b(?:trf\.?|mb\s*way)\s+para\b",                         # PT: Trf. Mb Way Para
            r"\btransfer(?:ência|encia)?\s+para\b",                     # PT: Transferência para
            r"\btransfer(?:encia)?\s+a\b",                              # ES: Transferencia a
            r"\bbizum\s+para\b",                                        # ES/PT: Bizum para
            r"\bpix\s+(?:enviado|saída|para)\b",                       # BR: Pix enviado
            r"\btransfer\s+to\b",                                      # EN: Transfer to
            r"\bsent\s+to\b",                                           # EN: Sent to
            r"\boutgoing\s+transfer\b",                                 # EN: Outgoing transfer
            r"\bpayment\s+to\b",                                        # EN: Payment to
            r"\bvirement\s+vers\b",                                     # FR: Virement vers
            r"\btransfert\s+vers\b",                                    # FR: Transfert vers
        ]
        trf_enviada = re.compile("|".join(f"(?:{p})" for p in trf_enviada_patterns), re.IGNORECASE)
        # Descrição genérica = sem destinatário/comerciante (ex.: "Trf. Mb Way Para" só). Se tiver "Para McDonald's", não é genérica.
        def _transfer_description_is_generic(description: str, is_enviada: bool) -> bool:
            if not description or len(description) < 10:
                return True
            # Remove o prefixo até (e incluindo) "para"/"to"/"vers" ou "de"/"from"; o que sobra é o destinatário/origem
            if is_enviada:
                m = re.search(r"(?:para|to|vers|enviado|saída)\s*", description, re.IGNORECASE)
            else:
                m = re.search(r"(?:de|from|reçu)\s*", description, re.IGNORECASE)
            if not m:
                return True
            rest = description[m.end() :].strip()
            if len(rest) <= 3:
                return True
            if re.search(r"[a-zA-ZÀ-ÿ]{4,}", rest):
                return False
            return True

        # Meses PT para parse de data "06 fev 2026"
        _meses_pt = {"jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6, "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12}

        def parse_date_from_doc(s: str):
            if not s or len(s) < 6:
                return today_str
            s = s.strip()[:50]
            # YYYY-MM-DD
            m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
            if m:
                return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
            # DD mmm YYYY (06 fev 2026)
            m = re.match(r"(\d{1,2})\s+(\w{3})\s+(\d{4})", s, re.IGNORECASE)
            if m:
                mes = _meses_pt.get(m.group(2).lower()[:3])
                if mes:
                    return f"{m.group(3)}-{mes:02d}-{int(m.group(1)):02d}"
            # DD/MM/YYYY
            m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
            if m:
                return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
            return today_str

        for item in items:
            try:
                amount = float(item.get("amount", 0))
            except (TypeError, ValueError):
                continue
            if amount <= 0:
                continue
            desc = (item.get("description") or "").strip()[:255] or "Movimento conta"
            tipo = (item.get("type") or "expense").lower()
            if tipo not in ("expense", "income"):
                tipo = "expense"
            date_str = parse_date_from_doc((item.get("date") or today_str).strip())
            cat_name = (item.get("category") or "").strip()

            # Forçar tipo e categoria para transferências só quando a descrição é genérica (sem destinatário/comerciante)
            if trf_recebida.search(desc):
                tipo = "income"
                if _transfer_description_is_generic(desc, False):
                    cat_name = first_income_cat if income_names else (cat_name or "Outros")
                elif not cat_name or cat_name not in (income_names or []):
                    cat_name = first_income_cat if income_names else (cat_name or "Outros")
            elif trf_enviada.search(desc):
                tipo = "expense"
                desc_generic = _transfer_description_is_generic(desc, True)
                # Só forçar Despesas gerais se: descrição genérica OU IA não devolveu categoria específica válida
                if desc_generic:
                    if not cat_name or _aliment.search(cat_name or ""):
                        cat_name = transfer_expense_cat
                else:
                    # Descrição com destinatário (ex.: "Para McDonald's") — manter categoria da IA se for válida (ex.: Alimentação)
                    if not cat_name or cat_name not in (expense_names or []):
                        cat_name = transfer_expense_cat

            if not cat_name and default_cat_name:
                cat_name = default_cat_name
            out.append({
                "amount": amount,
                "description": desc,
                "type": tipo,
                "date": date_str,
                "category": cat_name or None,
            })
        if not out:
            return None
        logger.info("[parse_document_with_openai] Extraídas %s transações do ficheiro", len(out))
        return {"transactions": out}
    except OpenAIRateLimitError:
        raise
    except Exception as e:
        logger.warning("[parse_document_with_openai] Erro: %s", e)
        return None


def _build_parsed_from_photo_result(
    photo_result: Dict, workspace: models.Workspace, db: Session
) -> Optional[Dict]:
    """
    Converte o resultado de process_photo_with_openai (uma transação ou lista de transações)
    no mesmo formato que parse_transaction: um dict único ou {"multiple": True, "transactions": [...]}.
    Otimização: carrega categorias uma vez por batch e reutiliza em todas as transações.
    """
    if not photo_result:
        return None
    if photo_result.get("transactions") and isinstance(photo_result["transactions"], list):
        raw_count = len(photo_result["transactions"])
        # Carregar categorias uma vez para todo o workspace (evita N queries iguais)
        all_cats = db.query(models.Category).filter(
            models.Category.workspace_id == workspace.id
        ).all()
        categories_by_type = {
            "expense": [c for c in all_cats if c.type == "expense"],
            "income": [c for c in all_cats if c.type == "income"],
        }
        transactions = []
        # A partir de 4 transações, processar em paralelo (categorização IA é o gargalo)
        if raw_count >= 4:
            def _cat_ref(c):
                return SimpleNamespace(
                    id=c.id, name=c.name, type=c.type,
                    vault_type=getattr(c, "vault_type", "none"),
                )
            refs = {
                t: [_cat_ref(c) for c in cats]
                for t, cats in categories_by_type.items()
            }

            def _process_one(item):
                session = SessionLocal()
                try:
                    return _parsed_from_photo(
                        item, workspace.id, session, categories_by_type=refs
                    )
                finally:
                    session.close()

            with ThreadPoolExecutor(max_workers=4) as executor:
                results = list(executor.map(_process_one, photo_result["transactions"]))
            transactions = [r for r in results if r]
        else:
            for item in photo_result["transactions"]:
                parsed_one = _parsed_from_photo(
                    item, workspace, db, categories_by_type=categories_by_type
                )
                if parsed_one:
                    transactions.append(parsed_one)
        if not transactions:
            logger.warning("[_build_parsed_from_photo_result] Nenhuma transação válida em %s itens", raw_count)
            return None
        ok_count = len(transactions)
        if ok_count < raw_count:
            logger.info("[_build_parsed_from_photo_result] %s transações OK, %s ignoradas (dados inválidos)", ok_count, raw_count - ok_count)
        else:
            logger.info("[_build_parsed_from_photo_result] %s transações validadas com sucesso", ok_count)
        if len(transactions) == 1:
            return transactions[0]
        return {"multiple": True, "transactions": transactions}
    return _parsed_from_photo(photo_result, workspace, db)


def _parsed_from_photo(
    photo_data: Dict,
    workspace: models.Workspace,
    db: Session,
    *,
    categories_by_type: Optional[Dict[str, List[models.Category]]] = None,
) -> Optional[Dict]:
    """
    Constrói um dict 'parsed' (mesmo formato que parse_transaction para transação única)
    a partir de um item (amount, description, type, date, category) devolvido pela Vision.
    Se categories_by_type for passado, evita query de categorias e reutiliza a lista para category_obj.
    workspace pode ser objeto Workspace ou UUID (path paralelo).
    """
    workspace_id = getattr(workspace, "id", None) or workspace
    description = (photo_data.get("description") or "").strip()[:255]
    try:
        amount = float(photo_data.get("amount", 0))
    except (TypeError, ValueError) as e:
        logger.warning("[_parsed_from_photo] amount inválido: %s", e)
        return None
    tipo = (photo_data.get("type") or "expense").lower()
    if tipo not in ("expense", "income"):
        tipo = "expense"
    if not description or amount <= 0:
        logger.warning("[_parsed_from_photo] Dados inválidos: description vazia ou amount<=0 (description=%r amount=%s)", description or "(vazio)", amount)
        return None
    date_str = photo_data.get("date") or ""
    transaction_date = date.today()
    if date_str:
        try:
            transaction_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            pass
    if categories_by_type is not None:
        categories = categories_by_type.get(tipo) or []
    else:
        categories = db.query(models.Category).filter(
            models.Category.workspace_id == workspace_id,
            models.Category.type == tipo,
        ).all()
    if not categories:
        logger.warning("[_parsed_from_photo] Sem categorias para workspace_id=%s tipo=%s", workspace_id, tipo)
        return None

    # Se a Vision devolveu uma categoria, tentar usar (match exato ou por similaridade)
    category_id = None
    suggested_category_name = None
    inference_source = "legacy_fallback"
    decision_reason = "legacy_fallback"
    needs_review = False
    vision_category_name = (photo_data.get("category") or "").strip()
    if vision_category_name:
        exact = next((c for c in categories if c.name.strip() == vision_category_name), None)
        if exact:
            category_id = exact.id
            inference_source = "openai_vision"
            decision_reason = "openai_vision"
        else:
            match = find_best_category_match(vision_category_name, categories, threshold=0.7)
            if match:
                category_id = match.id
                inference_source = "openai_vision"
                decision_reason = "openai_vision"
            else:
                suggested_category_name = vision_category_name[:100]

    if not category_id:
        try:
            from ..core.categorization_engine import infer_category
            from ..core.config import settings as _settings
            cat_id, source, needs_review, _conf, reason, _explain = infer_category(
                description,
                workspace_id,
                tipo,
                categories,
                db,
                models,
                _settings,
                explicit_category_id=None,
                use_gemini=True,
            )
            category_id = cat_id
            inference_source = source
            decision_reason = reason
        except Exception as e:
            logger.warning("Categorização falhou para foto: %s", e)
            _generic = {'despesas gerais', 'general expenses', 'dépenses générales', 'outros', 'other'}
            _specific = [c for c in categories if c.name.lower() not in _generic]
            category_id = (_specific[0] if _specific else categories[0]).id if categories else None
    if not category_id and categories and not suggested_category_name:
        _generic = {'despesas gerais', 'general expenses', 'dépenses générales', 'outros', 'other'}
        _specific = [c for c in categories if c.name.lower() not in _generic]
        category_id = (_specific[0] if _specific else categories[0]).id
    # Obter category_obj da lista já carregada (evita query extra)
    category_obj = next((c for c in categories if c.id == category_id), None) if category_id else None
    is_vault_category = category_obj and getattr(category_obj, "vault_type", "none") != "none"
    result = {
        "amount": amount,
        "description": description,
        "type": tipo,
        "category_id": category_id,
        "inference_source": inference_source,
        "decision_reason": decision_reason,
        "needs_review": needs_review,
        "is_vault": is_vault_category,
        "is_vault_withdrawal": False,
        "transaction_date": transaction_date,
        "suggested_category_name": suggested_category_name[:100] if suggested_category_name else None,
    }
    return result


# Redimensionar/comprimir imagem para Vision: menos tokens = resposta mais rápida
VISION_MAX_PIXELS = 1024
VISION_JPEG_QUALITY = 82
VISION_MIN_BYTES_TO_COMPRESS = 80 * 1024  # Só comprimir se > ~80KB


def _compress_image_for_vision(content: bytes, file_path: str, content_len: int):
    """
    Redimensiona e comprime a imagem para reduzir payload e tokens na Vision.
    Retorna (bytes_finais, mime). Se falhar ou imagem já pequena, devolve original.
    """
    mime = "image/jpeg"
    if file_path and file_path.lower().endswith(".png"):
        mime = "image/png"
    elif file_path and file_path.lower().endswith(".webp"):
        mime = "image/webp"
    if content_len < VISION_MIN_BYTES_TO_COMPRESS:
        return content, mime
    try:
        from io import BytesIO
        from PIL import Image
        img = Image.open(BytesIO(content))
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        w, h = img.size
        if w <= VISION_MAX_PIXELS and h <= VISION_MAX_PIXELS and content_len < 200 * 1024:
            return content, mime
        ratio = min(VISION_MAX_PIXELS / w, VISION_MAX_PIXELS / h, 1.0)
        if ratio < 1.0:
            new_w, new_h = int(w * ratio), int(h * ratio)
            img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
        buf = BytesIO()
        img.save(buf, format="JPEG", quality=VISION_JPEG_QUALITY, optimize=True)
        out = buf.getvalue()
        logger.info("[OpenAI Vision] Imagem comprimida: %s -> %s bytes (%.0f%%)", content_len, len(out), 100 * len(out) / max(1, content_len))
        return out, "image/jpeg"
    except Exception as e:
        logger.warning("[OpenAI Vision] Compressão falhou, usa original: %s", e)
        return content, mime


def _download_telegram_file(file_id: str, max_size_mb: int = 20) -> Optional[bytes]:
    """Descarrega um ficheiro do Telegram por file_id. Devolve o conteúdo em bytes ou None."""
    if not settings.TELEGRAM_BOT_TOKEN:
        return None
    try:
        get_file_url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/getFile?file_id={file_id}"
        r = requests.get(get_file_url, timeout=10)
        r.raise_for_status()
        data = r.json()
        if not data.get("ok") or "result" not in data:
            return None
        file_path = data["result"].get("file_path")
        if not file_path:
            return None
        download_url = f"https://api.telegram.org/file/bot{settings.TELEGRAM_BOT_TOKEN}/{file_path}"
        resp = requests.get(download_url, timeout=30)
        resp.raise_for_status()
        content = resp.content
        if not content or len(content) > max_size_mb * 1024 * 1024:
            return None
        return content
    except Exception as e:
        logger.warning("[Telegram] Download de ficheiro falhou: %s", e)
        return None


def _extract_text_from_pdf(content: bytes) -> Optional[str]:
    """Extrai texto de um PDF. Devolve string com newlines ou None se falhar."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(content))
        parts = []
        for page in reader.pages:
            try:
                t = page.extract_text()
                if t and t.strip():
                    parts.append(t.strip())
            except Exception:
                continue
        if not parts:
            return None
        return "\n".join(parts)
    except Exception as e:
        logger.warning("[Telegram] Extração de texto do PDF falhou: %s", e)
        return None


def _extract_text_from_csv(content: bytes) -> Optional[str]:
    """Extrai linhas tipo 'descrição valor€' de CSV (; ou ,). Devolve texto para parse_transaction."""
    import csv
    raw = None
    for enc in ('utf-8', 'utf-8-sig', 'latin-1', 'cp1252'):
        try:
            raw = content.decode(enc)
            break
        except Exception:
            continue
    if not raw or not raw.strip():
        return None
    lines = raw.strip().splitlines()
    if not lines:
        return None
    # Detetar delimitador (; ou , ou tab)
    first = lines[0]
    delim = ';' if ';' in first and first.count(';') >= 1 else (',' if ',' in first else '\t')
    reader = csv.reader(io.StringIO(raw), delimiter=delim)
    rows = list(reader)
    if not rows:
        return None
    # Primeira linha pode ser cabeçalho: se segunda linha tiver número, primeira é header
    start = 0
    if len(rows) > 1:
        second = rows[1]
        has_num = any(re.search(r'-?\d+[.,]\d*', str(c)) for c in second)
        if has_num and not any(re.search(r'-?\d+[.,]\d*', str(c)) for c in rows[0]):
            start = 1
    amount_pattern = re.compile(r'(-?\d{1,3}(?:[.\s]\d{3})*(?:[.,]\d+)?)\s*(?:€|eur|euros?|e)?\s*$', re.IGNORECASE)
    parts = []
    for row in rows[start:]:
        if not row:
            continue
        row = [str(c).strip() for c in row if c is not None]
        amount_val = None
        desc_parts = []
        for i, cell in enumerate(row):
            if not cell:
                continue
            m = amount_pattern.search(cell.replace(' ', ''))
            if m:
                try:
                    v = float(m.group(1).replace(' ', '').replace('.', '').replace(',', '.'))
                    if abs(v) < 1e9:
                        amount_val = v
                        desc_parts = [c for j, c in enumerate(row) if j != i and c and not amount_pattern.search(str(c).replace(' ', ''))]
                        if not desc_parts:
                            desc_parts = [c for j, c in enumerate(row) if j != i and c]
                        break
                except ValueError:
                    pass
            desc_parts.append(cell)
        if amount_val is not None and (desc_parts or row):
            desc = ' '.join(desc_parts) if desc_parts else (row[0] if row else 'Movimento')
            desc = re.sub(r'\s+', ' ', desc).strip()[:200]
            if desc:
                parts.append(f"{desc} {amount_val}€")
    if not parts:
        return None
    return ' '.join(parts)


def _extract_text_from_xlsx(content: bytes) -> Optional[str]:
    """Extrai linhas tipo 'descrição valor€' de Excel (.xlsx)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            return None
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        logger.warning("[Telegram] Extração Excel falhou: %s", e)
        return None
    if not rows:
        return None
    start = 0
    if len(rows) > 1:
        second = rows[1]
        has_num = any(v is not None and re.search(r'-?\d+[.,]\d*', str(v)) for v in (second or []))
        first_vals = rows[0] or []
        if has_num and not any(v is not None and re.search(r'-?\d+[.,]\d*', str(v)) for v in first_vals):
            start = 1
    amount_pattern = re.compile(r'(-?\d{1,3}(?:[.\s]\d{3})*(?:[.,]\d+)?)\s*(?:€|eur|euros?|e)?\s*$', re.IGNORECASE)
    parts = []
    for row in rows[start:]:
        if not row:
            continue
        row = [str(v).strip() if v is not None else '' for v in row]
        amount_val = None
        desc_parts = []
        for i, cell in enumerate(row):
            if not cell:
                continue
            cell_clean = cell.replace(' ', '')
            m = amount_pattern.search(cell_clean)
            if m:
                try:
                    v = float(m.group(1).replace(' ', '').replace('.', '').replace(',', '.'))
                    if abs(v) < 1e9:
                        amount_val = v
                        desc_parts = [c for j, c in enumerate(row) if j != i and c and not amount_pattern.search(str(c).replace(' ', ''))]
                        if not desc_parts:
                            desc_parts = [c for j, c in enumerate(row) if j != i and c]
                        break
                except ValueError:
                    pass
            desc_parts.append(cell)
        if amount_val is not None and (desc_parts or any(row)):
            desc = ' '.join(desc_parts) if desc_parts else (row[0] or 'Movimento')
            desc = re.sub(r'\s+', ' ', desc).strip()[:200]
            if desc:
                parts.append(f"{desc} {amount_val}€")
    if not parts:
        return None
    return ' '.join(parts)


def process_photo_with_openai(file_id: str, categories: List[models.Category]) -> Optional[Dict]:
    """
    Descarrega a foto do Telegram, comprime se necessário, e envia para OpenAI vision
    para extrair transações (uma ou lista para extratos).
    """
    logger.info("[OpenAI Vision] Início process_photo_with_openai file_id=%s categories_count=%s", file_id[:20] if file_id else None, len(categories) if categories else 0)
    if not settings.TELEGRAM_BOT_TOKEN:
        logger.warning("[OpenAI Vision] TELEGRAM_BOT_TOKEN não configurado")
        return None
    if not settings.OPENAI_API_KEY:
        logger.warning("[OpenAI Vision] OPENAI_API_KEY não configurado")
        return None
    try:
        # 1. Obter file_path do Telegram
        get_file_url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/getFile?file_id={file_id}"
        r = requests.get(get_file_url, timeout=10)
        r.raise_for_status()
        data = r.json()
        if not data.get("ok") or "result" not in data:
            logger.warning("[OpenAI Vision] Telegram getFile falhou: ok=%s result_present=%s body=%s", data.get("ok"), "result" in data, data)
            return None
        file_path = data["result"].get("file_path")
        if not file_path:
            logger.warning("[OpenAI Vision] getFile sem file_path: %s", data.get("result"))
            return None
        logger.info("[OpenAI Vision] file_path obtido: %s", file_path)
        # 2. Descarregar o ficheiro
        download_url = f"https://api.telegram.org/file/bot{settings.TELEGRAM_BOT_TOKEN}/{file_path}"
        img_resp = requests.get(download_url, timeout=15)
        img_resp.raise_for_status()
        content = img_resp.content
        content_len = len(content) if content else 0
        if not content:
            logger.warning("[OpenAI Vision] Download vazio")
            return None
        if content_len > 20 * 1024 * 1024:
            logger.warning("[OpenAI Vision] Imagem demasiado grande: %s bytes (máx 20MB)", content_len)
            return None
        logger.info("[OpenAI Vision] Imagem descarregada: %s bytes", content_len)
        content, mime = _compress_image_for_vision(content, file_path, content_len)
        import base64
        from openai import OpenAI
        from datetime import datetime as dt
        b64 = base64.b64encode(content).decode("utf-8")
        data_url = f"data:{mime};base64,{b64}"
        logger.info("[OpenAI Vision] Payload após compressão: %s bytes", len(content))

        expense_names = [c.name for c in categories if getattr(c, "type", "expense") == "expense"]
        income_names = [c.name for c in categories if getattr(c, "type", "income") == "income"]
        cats_expense = ", ".join(expense_names) if expense_names else "(nenhuma)"
        cats_income = ", ".join(income_names) if income_names else "(nenhuma)"

        prompt = f"""Extrai transações desta imagem.

Se for EXTRATO/TABELA (Data, Descrição, Valor, Categoria, Comerciante): lista TODAS as linhas. Cada linha = uma transação. Data→YYYY-MM-DD. Valor negativo→amount positivo, type expense. Categoria na tabela ou infere. Ignora ID e Saldo.
Se for RECIBO único: uma transação.

Data hoje: {dt.now().strftime('%Y-%m-%d')}
Categorias (nome EXATO): expense={cats_expense} | income={cats_income}

JSON só, sem markdown:
{{"transactions":[{{"amount":n,"description":"...","type":"expense","date":"YYYY-MM-DD","category":"NomeExato"}}]}}"""
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
        logger.info("[OpenAI Vision] A chamar OpenAI (imagem %s bytes)...", len(content))
        VISION_RETRY_DELAYS = [2, 5]
        text_response = ""
        for attempt in range(1 + len(VISION_RETRY_DELAYS)):
            try:
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt},
                                {"type": "image_url", "image_url": {"url": data_url}},
                            ],
                        }
                    ],
                    max_tokens=1500,
                    temperature=0.05,
                )
                if response.choices and response.choices[0].message.content:
                    text_response = response.choices[0].message.content.strip()
                break
            except OpenAIRateLimitError:
                raise
            except Exception as api_err:
                err_str = str(api_err).lower()
                is_429 = (
                    getattr(api_err, "status_code", None) == 429
                    or "429" in str(api_err)
                    or "rate_limit" in err_str
                    or "too many requests" in err_str
                )
                is_timeout = "timeout" in err_str or "timed out" in err_str
                if (is_429 or is_timeout) and attempt < len(VISION_RETRY_DELAYS):
                    delay = VISION_RETRY_DELAYS[attempt]
                    logger.warning("[OpenAI Vision] Tentativa %s falhou (%s), retry em %ss", attempt + 1, api_err, delay)
                    time.sleep(delay)
                    continue
                if is_429:
                    raise OpenAIRateLimitError(api_err) from api_err
                raise
        if not text_response:
            logger.warning("[OpenAI Vision] OpenAI respondeu sem content. choices=%s", len(response.choices) if response.choices else 0)
            return None
        logger.info("[OpenAI Vision] Resposta bruta (primeiros 500 chars): %s", (text_response[:500] + "..." if len(text_response) > 500 else text_response))
        clean = re.search(r"\{[\s\S]*\}", text_response)
        if clean:
            text_response = clean.group(0)
        try:
            parsed = json.loads(text_response)
        except json.JSONDecodeError as je:
            logger.warning("[OpenAI Vision] JSON inválido: %s. raw=%s", je, text_response[:300])
            return None
        transactions_raw = parsed.get("transactions")
        if isinstance(transactions_raw, list) and len(transactions_raw) > 0:
            out = []
            for item in transactions_raw:
                amount = float(item.get("amount", 0))
                description = (item.get("description") or "").strip()[:255]
                tipo = (item.get("type") or "expense").lower()
                if tipo not in ("expense", "income"):
                    tipo = "expense"
                date_str = item.get("date") or dt.now().strftime("%Y-%m-%d")
                category_name = (item.get("category") or "").strip()
                if not description:
                    description = "Transação"
                if amount <= 0 and tipo == "expense":
                    amount = abs(amount)
                out.append({
                    "amount": amount,
                    "description": description,
                    "type": tipo,
                    "date": date_str,
                    "category": category_name or None,
                })
            logger.info("[OpenAI Vision] Extraídas %s transações da imagem", len(out))
            return {"transactions": out}
        if isinstance(parsed.get("amount"), (int, float)):
            amount = float(parsed.get("amount", 0))
            description = (parsed.get("description") or "").strip()[:255]
            tipo = (parsed.get("type") or "expense").lower()
            if tipo not in ("expense", "income"):
                tipo = "expense"
            date_str = parsed.get("date") or dt.now().strftime("%Y-%m-%d")
            category_name = (parsed.get("category") or "").strip()
            if not description or amount <= 0:
                logger.warning("[OpenAI Vision] Dados inválidos: description=%r amount=%s", description or "(vazio)", amount)
                return None
            return {
                "amount": amount,
                "description": description,
                "type": tipo,
                "date": date_str,
                "category": category_name or None,
            }
        logger.warning("[OpenAI Vision] Resposta sem 'transactions' nem campos de uma transação: %s", list(parsed.keys()) if isinstance(parsed, dict) else type(parsed))
        return None
    except OpenAIRateLimitError:
        raise
    except Exception as e:
        err_str = str(e).lower()
        is_rate_limit = (
            getattr(e, "status_code", None) == 429
            or "429" in str(e)
            or "rate_limit" in err_str
            or "too many requests" in err_str
            or "quota" in err_str
        )
        if is_rate_limit:
            logger.warning("[OpenAI Vision] Rate limit (429) - demasiados pedidos ou quota excedida: %s", e)
            raise OpenAIRateLimitError(e) from e
        logger.exception("[OpenAI Vision] Erro ao processar foto: %s", e)
        return None


def _ogg_to_mp3_bytes(content: bytes) -> Optional[bytes]:
    """Converte áudio OGG (ex.: mensagem de voz Telegram) para MP3 para compatibilidade com Whisper."""
    try:
        from pydub import AudioSegment
        seg = AudioSegment.from_file(io.BytesIO(content), format="ogg")
        buf = io.BytesIO()
        seg.export(buf, format="mp3")
        buf.seek(0)
        return buf.read()
    except Exception as e:
        logger.warning("[Whisper] Conversão OGG->MP3 falhou: %s", e)
        return None


def transcribe_audio_from_telegram(message: dict) -> Optional[str]:
    """
    Obtém file_id de message['voice'] ou message['audio'], descarrega o ficheiro,
    converte OGG para MP3 se necessário, e envia para OpenAI Whisper. Devolve o texto transcrito ou None.
    """
    voice = message.get("voice") or message.get("audio")
    if not voice:
        return None
    file_id = voice.get("file_id")
    if not file_id:
        return None
    if not settings.TELEGRAM_BOT_TOKEN or not settings.OPENAI_API_KEY:
        logger.warning("[Whisper] TELEGRAM_BOT_TOKEN ou OPENAI_API_KEY não configurado")
        return None
    try:
        get_file_url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/getFile?file_id={file_id}"
        r = requests.get(get_file_url, timeout=10)
        r.raise_for_status()
        data = r.json()
        if not data.get("ok") or "result" not in data:
            logger.warning("[Whisper] getFile falhou: %s", data)
            return None
        file_path = data["result"].get("file_path")
        if not file_path:
            return None
        download_url = f"https://api.telegram.org/file/bot{settings.TELEGRAM_BOT_TOKEN}/{file_path}"
        resp = requests.get(download_url, timeout=15)
        resp.raise_for_status()
        content = resp.content
        if not content or len(content) > 25 * 1024 * 1024:
            logger.warning("[Whisper] Áudio vazio ou >25MB: %s bytes", len(content) if content else 0)
            return None
        # Whisper suporta mp3, m4a, wav, webm; Telegram voice é normalmente .ogg
        ext = (file_path or "").lower().split(".")[-1] if "." in (file_path or "") else ""
        if ext == "ogg":
            mp3_content = _ogg_to_mp3_bytes(content)
            if mp3_content:
                content = mp3_content
                ext = "mp3"
            else:
                logger.warning("[Whisper] Áudio OGG sem conversão (instala ffmpeg para voz Telegram)")
                return None
        suffix = f".{ext}" if ext in ("mp3", "mp4", "mpeg", "mpga", "m4a", "wav", "webm") else ".mp3"
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=True) as tmp:
            tmp.write(content)
            tmp.flush()
            from openai import OpenAI
            client = OpenAI(api_key=settings.OPENAI_API_KEY)
            with open(tmp.name, "rb") as f:
                transcript = client.audio.transcriptions.create(model="whisper-1", file=f)
        text = (transcript.text or "").strip()
        if not text:
            return None
        logger.info("[Whisper] Transcrição: %s...", text[:80] if len(text) > 80 else text)
        return text
    except OpenAIRateLimitError:
        raise
    except Exception as e:
        err_str = str(e).lower()
        if "429" in err_str or "rate" in err_str or "quota" in err_str:
            raise OpenAIRateLimitError(e) from e
        logger.exception("[Whisper] Erro ao transcrever áudio: %s", e)
        return None


def _chunk_text_for_telegram(text: str, max_len: int = TELEGRAM_MAX_MESSAGE_LENGTH) -> List[str]:
    """Divide texto em blocos <= max_len, cortando preferencialmente em newline."""
    if len(text) <= max_len:
        return [text] if text else []
    chunks = []
    rest = text
    while rest:
        if len(rest) <= max_len:
            chunks.append(rest)
            break
        block = rest[:max_len]
        last_nl = block.rfind("\n")
        if last_nl > max_len // 2:
            cut = last_nl + 1
        else:
            cut = max_len
        chunks.append(rest[:cut])
        rest = rest[cut:].lstrip("\n")
    return chunks


def send_telegram_msg(chat_id: int, text: str, reply_markup: Optional[Dict] = None, pin_message: bool = False):
    """Envia mensagem para o Telegram. Mensagens longas são divididas em várias (limite 4096 carateres)."""
    if not settings.TELEGRAM_BOT_TOKEN:
        logger.warning("TELEGRAM_BOT_TOKEN não configurado")
        return None

    if not (text or text.strip()):
        return None

    url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/sendMessage"
    chunks = _chunk_text_for_telegram(text.strip())
    last_result = None

    for i, chunk in enumerate(chunks):
        use_html = i == 0 and len(chunks) == 1
        # Telegram HTML parse_mode não suporta <br>; usar \n para quebras de linha
        if use_html:
            chunk = chunk.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
        payload = {
            "chat_id": chat_id,
            "text": chunk,
        }
        if use_html:
            payload["parse_mode"] = "HTML"
        if reply_markup and i == len(chunks) - 1:
            payload["reply_markup"] = reply_markup

        try:
            response = requests.post(url, json=payload, timeout=5)
            response.raise_for_status()
            last_result = response.json()
        except requests.exceptions.HTTPError as e:
            if getattr(e, "response", None) and getattr(e.response, "status_code", None) == 400:
                try:
                    err_body = e.response.json() if e.response.text else {}
                    err_desc = (err_body.get("description") or "").lower()
                except Exception:
                    err_desc = ""
                if "too long" in err_desc or "message is too long" in err_desc:
                    logger.warning("Mensagem ainda longa após chunking, a truncar: %s", err_desc)
                    payload["text"] = chunk[:TELEGRAM_MAX_MESSAGE_LENGTH - 50] + "\n\n(… truncado)"
                if "parse_mode" in payload:
                    payload.pop("parse_mode", None)
                try:
                    response = requests.post(url, json=payload, timeout=5)
                    response.raise_for_status()
                    last_result = response.json()
                except Exception as e2:
                    logger.error("Erro ao enviar mensagem Telegram (sem parse_mode): %s", e2)
                    return last_result
            else:
                logger.error("Erro HTTP ao enviar mensagem Telegram: %s - %s", getattr(e.response, "status_code", ""), getattr(e.response, "text", ""))
                return last_result
        except Exception as e:
            logger.error("Erro ao enviar mensagem Telegram: %s", e)
            return last_result

    sent_message_id = None
    if last_result:
        sent_message_id = (last_result.get("result") or {}).get("message_id")

    if last_result and pin_message and len(chunks) == 1:
        try:
            message_id = sent_message_id
            if message_id:
                pin_url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/pinChatMessage"
                requests.post(
                    pin_url,
                    json={"chat_id": chat_id, "message_id": message_id, "disable_notification": True},
                    timeout=5,
                )
                logger.info("Mensagem fixada: message_id=%s", message_id)
        except Exception as e:
            logger.warning("Erro ao fixar mensagem: %s", e)

    return sent_message_id


def _send_typing_action(chat_id: int) -> bool:
    """Envia indicador 'a escrever...' (três pontos) no Telegram."""
    if not settings.TELEGRAM_BOT_TOKEN:
        return False
    url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/sendChatAction"
    try:
        r = requests.post(url, json={"chat_id": chat_id, "action": "typing"}, timeout=5)
        return r.ok
    except Exception as e:
        logger.warning("Erro ao enviar typing action: %s", e)
        return False


def _delete_telegram_msg(chat_id: int, message_id: int) -> bool:
    """Apaga uma mensagem no Telegram (ex: indicador 'a pensar...')."""
    if not settings.TELEGRAM_BOT_TOKEN or not message_id:
        return False
    url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/deleteMessage"
    try:
        r = requests.post(url, json={"chat_id": chat_id, "message_id": message_id}, timeout=5)
        return r.ok
    except Exception as e:
        logger.warning("Erro ao apagar mensagem Telegram: %s", e)
        return False


def edit_telegram_message(chat_id: int, message_id: int, text: str, reply_markup: Optional[Dict] = None) -> bool:
    """Edita texto e/ou teclado de uma mensagem existente. Trunca se exceder o limite do Telegram."""
    if not settings.TELEGRAM_BOT_TOKEN:
        return False
    if len(text) > TELEGRAM_MAX_MESSAGE_LENGTH:
        text = text[: TELEGRAM_MAX_MESSAGE_LENGTH - 30] + "\n\n(… truncado)"
    # Telegram HTML não suporta <br>; usar \n
    text = text.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
    url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/editMessageText"
    payload = {"chat_id": chat_id, "message_id": message_id, "text": text, "parse_mode": "HTML"}
    if reply_markup:
        payload["reply_markup"] = reply_markup
    try:
        r = requests.post(url, json=payload, timeout=5)
        r.raise_for_status()
        return True
    except Exception as e:
        logger.warning("[Telegram] editMessageText falhou: %s", e)
        return False


def _build_batch_message_and_keyboard(
    chat_id: str,
    batch_id_hex: str,
    db: Session,
    t,
) -> tuple:
    """Constrói texto e inline_keyboard para a lista de pendentes do batch. Devolve (message_text, reply_markup) ou (None, None) se não houver pendentes."""
    pendents_batch = (
        db.query(models.TelegramPendingTransaction)
        .filter(
            models.TelegramPendingTransaction.chat_id == chat_id,
            models.TelegramPendingTransaction.batch_id.isnot(None),
        )
        .order_by(models.TelegramPendingTransaction.created_at)
        .all()
    )
    pendents_batch = [p for p in pendents_batch if p.batch_id and p.batch_id.hex[:16] == batch_id_hex]
    if not pendents_batch:
        return None, None
    lines = []
    total_cents = 0
    for p in pendents_batch:
        total_cents += p.amount_cents
        cat = db.query(models.Category).filter(models.Category.id == p.category_id).first()
        cat_name = cat.name if cat else "Outros"
        desc_display = _html_escape(_shorten_description_for_list(p.description))
        amount_display = "{:.2f}".format(abs(p.amount_cents) / 100).replace(".", ",")
        lines.append(t('list_pending_line').format(
            description=desc_display,
            amount=amount_display,
            category=cat_name,
        ))
    total_euros = abs(total_cents) / 100
    total_display = "{:.2f}".format(total_euros).replace(".", ",")
    message_text = (
        t('list_pending_header')
        + "".join(lines)
        + t('list_pending_total').format(total=total_display)
        + t('list_confirm_question')
    )
    # Apenas Confirmar tudo e Cancelar tudo (sem botões por linha)
    keyboard = [[
        {"text": t('button_confirm_all'), "callback_data": f"confirm_batch_{batch_id_hex}"},
        {"text": t('button_cancel_all'), "callback_data": f"cancel_batch_{batch_id_hex}"},
    ]]
    reply_markup = {"inline_keyboard": keyboard}
    return message_text, reply_markup


def setup_bot_commands():
    """Configura os comandos do bot no Telegram (aparecem no menu azul ao digitar /)"""
    if not settings.TELEGRAM_BOT_TOKEN:
        logger.warning("TELEGRAM_BOT_TOKEN não configurado - não é possível configurar comandos")
        return
    
    # Lista completa para aparecer no menu de comandos (max 100; descrição max 256 chars)
    commands = [
        {"command": "start", "description": "Iniciar e associar conta"},
        {"command": "info", "description": "📖 Guia e exemplos de formato"},
        {"command": "help", "description": "❓ Ajuda e comandos"},
        {"command": "ajuda", "description": "❓ Ajuda (atalho)"},
        {"command": "resumo", "description": "📊 Resumo do dia (hoje)"},
        {"command": "hoje", "description": "📊 Resumo do dia"},
        {"command": "mes", "description": "📊 Resumo do mês"},
        {"command": "pendentes", "description": "📋 Listar transações pendentes"},
        {"command": "clear", "description": "🧹 Limpar todas as pendentes"},
        {"command": "revoke", "description": "🔓 Desvincular Telegram da conta"},
        {"command": "idioma", "description": "🌐 Mudar idioma (pt / en)"},
        {"command": "categoria", "description": "🏷️ Categoria por defeito (nome ou stop)"},
    ]
    
    url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/setMyCommands"
    try:
        response = requests.post(url, json={"commands": commands}, timeout=5)
        response.raise_for_status()
        logger.info("Comandos do bot configurados com sucesso (menu /)")
    except Exception as e:
        logger.error("Erro ao configurar comandos do bot: %s", e)

def setup_bot_info():
    """Configura informações adicionais do bot (descrição, about, etc.)"""
    if not settings.TELEGRAM_BOT_TOKEN:
        logger.warning("TELEGRAM_BOT_TOKEN não configurado - não é possível configurar informações")
        return
    
    base_url = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}"
    
    # Configurar descrição curta (aparece no perfil do bot)
    try:
        short_desc = "🧘‍♂️ O teu ecossistema financeiro inteligente. Regista transações em segundos."
        requests.post(
            f"{base_url}/setMyShortDescription",
            json={'short_description': short_desc},
            timeout=5
        )
        logger.info("Descrição curta do bot configurada")
    except Exception as e:
        logger.warning(f"Erro ao configurar descrição curta: {str(e)}")
    
    # Configurar descrição completa (about)
    try:
        full_desc = (
            "Finly Bot\n\n"
            "💎 Regista transações financeiras rapidamente através do Telegram.\n\n"
            "🎯 Funcionalidades:\n"
            "• Categorização automática com IA\n"
            "• Suporte a múltiplas transações\n"
            "• Especifica categoria: Descrição - Categoria Valor€\n"
            "• Confirmação opcional de transações\n\n"
            "🧘‍♂️ Domina o teu dinheiro com simplicidade."
        )
        requests.post(
            f"{base_url}/setMyDescription",
            json={'description': full_desc},
            timeout=5
        )
        logger.info("Descrição completa do bot configurada")
    except Exception as e:
        logger.warning(f"Erro ao configurar descrição completa: {str(e)}")
    
    # Configurar nome do bot (se ainda não estiver configurado)
    try:
        bot_name = "Finly Bot"
        requests.post(
            f"{base_url}/setMyName",
            json={'name': bot_name},
            timeout=5
        )
        logger.info("Nome do bot configurado")
    except Exception as e:
        logger.warning(f"Erro ao configurar nome do bot: {str(e)}")

@router.post('/webhook')
@limiter.limit('30/minute')
async def telegram_webhook(
    request: Request, 
    db: Session = Depends(get_db),
    x_telegram_bot_api_secret_token: str | None = Header(None)
):
    """Webhook Telegram com validação de segurança"""
    logger.info("=" * 50)
    logger.info("Webhook Telegram recebido")
    logger.info(f"Headers: X-Telegram-Bot-Api-Secret-Token presente: {x_telegram_bot_api_secret_token is not None}")
    
    try:
        # Validação do secret token
        if settings.TELEGRAM_WEBHOOK_SECRET:
            logger.info(f"Validando secret token... (configurado: {bool(settings.TELEGRAM_WEBHOOK_SECRET)})")
            if not x_telegram_bot_api_secret_token or x_telegram_bot_api_secret_token != settings.TELEGRAM_WEBHOOK_SECRET:
                logger.warning(f"Tentativa de acesso ao webhook sem token válido. Recebido: {x_telegram_bot_api_secret_token is not None}, Esperado: {settings.TELEGRAM_WEBHOOK_SECRET[:10]}...")
                raise HTTPException(status_code=403, detail="Invalid secret token")
            logger.info("Secret token valido [OK]")
        else:
            logger.warning("TELEGRAM_WEBHOOK_SECRET não configurado - validação desativada")
        
        data = await request.json()
        logger.info(f"Payload recebido: {json.dumps(data, indent=2, ensure_ascii=False)[:500]}...")  # Primeiros 500 chars
        
        # Idempotência: ignorar update já processado (reenvios do Telegram)
        update_id = data.get('update_id')
        if _is_duplicate_update(update_id):
            logger.info("Update %s já processado (idempotência), ignorar", update_id)
            return {'status': 'duplicate'}
        
        # Processar callback_query (botões inline)
        if 'callback_query' in data:
            logger.info("Processando callback_query (botão inline)")
            callback_query = data['callback_query']
            chat_id = callback_query['message']['chat']['id']
            callback_data = callback_query.get('data', '')
            message_id = callback_query['message']['message_id']
            logger.info(f"Callback: chat_id={chat_id}, data={callback_data}")
            
            # Verificar rate limit
            if not check_rate_limit(str(chat_id)):
                user = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
                language = (user.language if user and user.language else None) or _telegram_lang(callback_query.get("from"))
                t = get_telegram_t(language)
                send_telegram_msg(chat_id, t('rate_limit'))
                return {'status': 'rate_limited'}
            
            # Buscar utilizador
            user = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user:
                lang_cb = _telegram_lang(callback_query.get("from"))
                t = get_telegram_t(lang_cb)
                send_telegram_msg(chat_id, t('session_expired'))
                return {'status': 'unauthorized'}
            
            # Obter linguagem do utilizador
            language = user.language if user.language else 'pt'
            t = get_telegram_t(language)
            
            # Processar callback
            if callback_data.startswith("confirm_batch_"):
                batch_id_hex = callback_data.replace("confirm_batch_", "").strip()[:16]
                logger.info("[Telegram] confirm_batch: batch_id_hex=%r chat_id=%s", batch_id_hex, chat_id)
                all_pending = db.query(models.TelegramPendingTransaction).filter(
                    models.TelegramPendingTransaction.chat_id == str(chat_id),
                ).all()
                # Match exato pelos 16 primeiros chars do batch_id (como no botão)
                batch_pendents = [p for p in all_pending if p.batch_id and p.batch_id.hex[:16] == batch_id_hex]
                logger.info("[Telegram] confirm_batch: all_pending=%s batch_pendents=%s", len(all_pending), len(batch_pendents))
                if not batch_id_hex or not batch_pendents:
                    send_telegram_msg(chat_id, t('transaction_not_found'))
                    try:
                        requests.post(
                            f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                            json={'callback_query_id': callback_query['id']},
                            timeout=5,
                        )
                    except Exception:
                        pass
                    return {'status': 'not_found'}
                for pending in batch_pendents:
                    transaction = models.Transaction(
                        workspace_id=pending.workspace_id,
                        category_id=pending.category_id,
                        amount_cents=pending.amount_cents,
                        description=pending.description,
                        inference_source=getattr(pending, 'inference_source', None),
                        decision_reason=getattr(pending, 'decision_reason', None),
                        needs_review=getattr(pending, 'needs_review', False),
                        transaction_date=getattr(pending, 'transaction_date', None) or date.today(),
                    )
                    db.add(transaction)
                    # Aprendizagem: guardar no cache para futuras mensagens (menos IA)
                    cache_key = _description_cache_key(pending.description)
                    if cache_key and pending.category_id:
                        category = db.query(models.Category).filter(models.Category.id == pending.category_id).first()
                        cat_name = category.name if category else "Outros"
                        tipo = "expense" if pending.amount_cents < 0 else "income"
                        save_cached_category(cache_key, pending.workspace_id, pending.category_id, cat_name, tipo, db, is_common=True)
                    db.delete(pending)
                db.commit()
                logger.info("[Telegram] confirm_batch: criadas %s transações para chat_id=%s", len(batch_pendents), chat_id)
                try:
                    requests.post(
                        f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                        json={'callback_query_id': callback_query['id']},
                        timeout=5,
                    )
                except Exception:
                    pass
                send_telegram_msg(chat_id, t('list_confirmed'))
                return {'status': 'confirmed'}
            elif callback_data.startswith("create_cat_"):
                # User escolheu "Sim, criar" categoria sugerida pela IA
                pending_id_hex = callback_data.replace("create_cat_", "")
                all_pending = db.query(models.TelegramPendingTransaction).filter(
                    models.TelegramPendingTransaction.chat_id == str(chat_id)
                ).all()
                pending = None
                for p in all_pending:
                    if p.id.hex[:16] == pending_id_hex and getattr(p, 'suggested_category_name', None):
                        pending = p
                        break
                if not pending:
                    send_telegram_msg(chat_id, t('transaction_not_found'))
                    try:
                        requests.post(
                            f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                            json={'callback_query_id': callback_query['id']},
                            timeout=5,
                        )
                    except Exception:
                        pass
                    return {'status': 'not_found'}
                suggested_name = (pending.suggested_category_name or "").strip()[:100]
                if not suggested_name:
                    send_telegram_msg(chat_id, t('transaction_not_found'))
                    return {'status': 'not_found'}
                tipo = "expense" if pending.amount_cents < 0 else "income"
                existing = db.query(models.Category).filter(
                    models.Category.workspace_id == pending.workspace_id,
                    func.lower(models.Category.name) == suggested_name.lower(),
                    models.Category.type == tipo,
                ).first()
                if existing:
                    new_category_id = existing.id
                    new_category_name = existing.name
                else:
                    new_cat = models.Category(
                        workspace_id=pending.workspace_id,
                        name=suggested_name,
                        type=tipo,
                        vault_type='none',
                        monthly_limit_cents=0,
                        color_hex='#3B82F6',
                        icon='Tag',
                        is_default=False,
                    )
                    db.add(new_cat)
                    db.flush()
                    new_category_id = new_cat.id
                    new_category_name = new_cat.name
                    logger.info("Categoria criada pelo user Telegram: %s (id=%s)", new_category_name, new_category_id)
                pending.category_id = new_category_id
                pending.suggested_category_name = None
                db.commit()
                try:
                    requests.post(
                        f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                        json={'callback_query_id': callback_query['id']},
                        timeout=5,
                    )
                except Exception:
                    pass
                msg = t('category_created_confirm').format(name=new_category_name)
                tipo_emoji = "" if pending.amount_cents < 0 else "💰"
                tipo_texto = t('type_expense') if pending.amount_cents < 0 else t('type_income')
                msg += "\n\n" + t('transaction_pending').format(
                    description=pending.description,
                    emoji=tipo_emoji,
                    amount=abs(pending.amount_cents) / 100,
                    category=new_category_name,
                    type=tipo_texto,
                    origin_line=_origin_line("openai", t),
                    date_line=_date_line(pending.transaction_date, t),
                )
                pending_id_hex_new = pending.id.hex[:16]
                reply_markup = {
                    "inline_keyboard": [[
                        {"text": t('button_confirm'), "callback_data": f"confirm_{pending_id_hex_new}"},
                        {"text": t('button_cancel'), "callback_data": f"cancel_{pending_id_hex_new}"},
                    ]]
                }
                send_telegram_msg(chat_id, msg, reply_markup)
                return {'status': 'category_created'}
            elif callback_data.startswith("skip_cat_"):
                # User escolheu "Não, cancelar" - apagar pendente
                pending_id_hex = callback_data.replace("skip_cat_", "")
                all_pending = db.query(models.TelegramPendingTransaction).filter(
                    models.TelegramPendingTransaction.chat_id == str(chat_id)
                ).all()
                pending = None
                for p in all_pending:
                    if p.id.hex[:16] == pending_id_hex and getattr(p, 'suggested_category_name', None):
                        pending = p
                        break
                if pending:
                    db.delete(pending)
                    db.commit()
                try:
                    requests.post(
                        f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                        json={'callback_query_id': callback_query['id']},
                        timeout=5,
                    )
                except Exception:
                    pass
                send_telegram_msg(chat_id, t('transaction_cancelled'))
                return {'status': 'cancelled'}
            elif callback_data.startswith("changecat_"):
                pending_id_hex = callback_data.replace("changecat_", "")
                all_p = db.query(models.TelegramPendingTransaction).filter(
                    models.TelegramPendingTransaction.chat_id == str(chat_id),
                ).all()
                pending = next((p for p in all_p if p.id.hex[:16] == pending_id_hex), None)
                if not pending or pending.batch_id:
                    try:
                        requests.post(
                            f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                            json={'callback_query_id': callback_query['id']},
                            timeout=5,
                        )
                    except Exception:
                        pass
                    return {'status': 'not_found'}
                tipo = "expense" if pending.amount_cents < 0 else "income"
                categories = db.query(models.Category).filter(
                    models.Category.workspace_id == pending.workspace_id,
                    models.Category.type == tipo,
                ).all()
                keyboard = []
                row = []
                for c in categories[:20]:
                    row.append({"text": (c.name[:20] + "…" if len(c.name) > 20 else c.name), "callback_data": f"setcat_{pending_id_hex}_{c.id}"})
                    if len(row) >= 2:
                        keyboard.append(row)
                        row = []
                if row:
                    keyboard.append(row)
                try:
                    requests.post(
                        f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                        json={'callback_query_id': callback_query['id']},
                        timeout=5,
                    )
                except Exception:
                    pass
                edit_telegram_message(
                    chat_id,
                    callback_query["message"]["message_id"],
                    t('change_category_prompt'),
                    {"inline_keyboard": keyboard},
                )
                return {'status': 'changecat_shown'}
            elif callback_data.startswith("setcat_"):
                parts = callback_data.replace("setcat_", "").split("_", 1)
                if len(parts) != 2:
                    return {'status': 'error'}
                pending_id_hex, category_id_str = parts[0], parts[1]
                all_p = db.query(models.TelegramPendingTransaction).filter(
                    models.TelegramPendingTransaction.chat_id == str(chat_id),
                ).all()
                pending = next((p for p in all_p if p.id.hex[:16] == pending_id_hex), None)
                if not pending or pending.batch_id:
                    try:
                        requests.post(
                            f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                            json={'callback_query_id': callback_query['id']},
                            timeout=5,
                        )
                    except Exception:
                        pass
                    return {'status': 'not_found'}
                try:
                    cat_uuid = uuid.UUID(category_id_str)
                except ValueError:
                    return {'status': 'error'}
                cat = db.query(models.Category).filter(
                    models.Category.id == cat_uuid,
                    models.Category.workspace_id == pending.workspace_id,
                ).first()
                if not cat:
                    return {'status': 'not_found'}
                pending.category_id = cat.id
                db.commit()
                # Aprender com a correção do user para melhorar categorizações futuras
                try:
                    from ..core.categorization_engine import learn_from_correction
                    tipo_lc = "expense" if pending.amount_cents < 0 else "income"
                    learn_from_correction(
                        pending.description or "",
                        cat.id, pending.workspace_id, tipo_lc, cat.name, db, models,
                    )
                except Exception as lc_err:
                    logger.warning("learn_from_correction falhou: %s", lc_err)
                try:
                    requests.post(
                        f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                        json={'callback_query_id': callback_query['id']},
                        timeout=5,
                    )
                except Exception:
                    pass
                tipo_emoji = "" if pending.amount_cents < 0 else "💰"
                tipo_texto = t('type_expense') if pending.amount_cents < 0 else t('type_income')
                msg = t('transaction_pending').format(
                    description=pending.description,
                    emoji=tipo_emoji,
                    amount=abs(pending.amount_cents) / 100,
                    category=cat.name,
                    type=tipo_texto,
                    origin_line=_origin_line(getattr(pending, 'inference_source', None), t),
                    date_line=_date_line(getattr(pending, 'transaction_date', None) or date.today(), t),
                )
                reply_markup = {
                    "inline_keyboard": [[
                        {"text": t('button_confirm'), "callback_data": f"confirm_{pending_id_hex}"},
                        {"text": t('button_cancel'), "callback_data": f"cancel_{pending_id_hex}"},
                    ]]
                }
                edit_telegram_message(chat_id, callback_query["message"]["message_id"], msg, reply_markup)
                return {'status': 'category_updated'}
            elif callback_data.startswith("cancel_batch_"):
                batch_id_hex = callback_data.replace("cancel_batch_", "")
                all_pending = db.query(models.TelegramPendingTransaction).filter(
                    models.TelegramPendingTransaction.chat_id == str(chat_id),
                ).all()
                batch_pendents = [p for p in all_pending if p.batch_id and p.batch_id.hex[:16] == batch_id_hex]
                for pending in batch_pendents:
                    db.delete(pending)
                db.commit()
                try:
                    requests.post(
                        f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                        json={'callback_query_id': callback_query['id']},
                        timeout=5,
                    )
                except Exception:
                    pass
                send_telegram_msg(chat_id, t('list_cancelled'))
                return {'status': 'cancelled'}
            elif callback_data.startswith("confirm_"):
                logger.info(f"Processando confirmacao de transacao: {callback_data}")
                # Confirmar transação
                pending_id_hex = callback_data.replace("confirm_", "")
                logger.info(f"Buscando pending transaction com hex: {pending_id_hex}")
                
                # Buscar por hex curto (primeiros 16 caracteres do UUID)
                # Buscar todas as transações pendentes deste chat e filtrar por UUID
                all_pending = db.query(models.TelegramPendingTransaction).filter(
                    models.TelegramPendingTransaction.chat_id == str(chat_id)
                ).all()
                logger.info(f"Encontradas {len(all_pending)} transacoes pendentes para chat_id={chat_id}")
                
                pending = None
                for p in all_pending:
                    logger.info(f"Comparando: {p.id.hex[:16]} com {pending_id_hex}")
                    if p.id.hex.startswith(pending_id_hex):
                        pending = p
                        logger.info(f"Match encontrado! Pending ID: {p.id}, workspace: {p.workspace_id}, amount: {p.amount_cents}")
                        break
                
                if not pending:
                    logger.warning(f"Pending transaction nao encontrada para hex: {pending_id_hex}")
                    send_telegram_msg(chat_id, t('transaction_not_found'))
                    return {'status': 'not_found'}
                
                from_batch = pending.batch_id is not None
                batch_id_hex = pending.batch_id.hex[:16] if pending.batch_id else None
                # Criar transação real
                logger.info(f"Criando transacao: workspace_id={pending.workspace_id}, category_id={pending.category_id}, amount_cents={pending.amount_cents}, description={pending.description}, transaction_date={pending.transaction_date}")
                transaction = models.Transaction(
                    workspace_id=pending.workspace_id,
                    category_id=pending.category_id,
                    amount_cents=pending.amount_cents,
                    description=pending.description,
                    inference_source=getattr(pending, 'inference_source', None),
                    decision_reason=getattr(pending, 'decision_reason', None),
                    needs_review=getattr(pending, 'needs_review', False),
                    transaction_date=pending.transaction_date
                )
                db.add(transaction)
                db.flush()
                logger.info(f"Transacao criada com ID: {transaction.id}, transaction_date: {transaction.transaction_date}, created_at: {transaction.created_at}")
                desc_for_msg = pending.description
                amount_cents_for_msg = pending.amount_cents
                category_id_for_msg = pending.category_id
                inference_source_for_msg = getattr(pending, 'inference_source', None)
                transaction_date_for_msg = getattr(pending, 'transaction_date', None) or date.today()
                cache_key = _description_cache_key(pending.description)
                if cache_key and pending.category_id:
                    category = db.query(models.Category).filter(models.Category.id == pending.category_id).first()
                    cat_name = category.name if category else "Outros"
                    tipo = "expense" if pending.amount_cents < 0 else "income"
                    save_cached_category(cache_key, pending.workspace_id, pending.category_id, cat_name, tipo, db, is_common=True)
                db.delete(pending)
                db.commit()
                logger.info("Transacao confirmada e commitada com sucesso")
                
                try:
                    requests.post(
                        f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                        json={'callback_query_id': callback_query['id']},
                        timeout=5,
                    )
                except Exception as e:
                    logger.error(f"Erro ao responder callback: {str(e)}")
                
                if from_batch and batch_id_hex:
                    message_id = callback_query.get("message", {}).get("message_id")
                    if message_id is not None:
                        msg_text, markup = _build_batch_message_and_keyboard(str(chat_id), batch_id_hex, db, t)
                        if msg_text and markup:
                            edit_telegram_message(chat_id, message_id, msg_text, markup)
                        else:
                            edit_telegram_message(chat_id, message_id, t('batch_list_empty'), {"inline_keyboard": []})
                else:
                    tipo_emoji = "" if amount_cents_for_msg < 0 else "💰"
                    tipo_texto = t('type_expense') if amount_cents_for_msg < 0 else t('type_income')
                    category = db.query(models.Category).filter(models.Category.id == category_id_for_msg).first()
                    category_name = category.name if category else "Outros"
                    origin_line = _origin_line(inference_source_for_msg, t)
                    send_telegram_msg(chat_id, t('transaction_confirmed').format(
                        description=desc_for_msg,
                        emoji=tipo_emoji,
                        amount=abs(amount_cents_for_msg)/100,
                        category=category_name,
                        type=tipo_texto,
                        origin_line=origin_line,
                        date_line=_date_line(transaction_date_for_msg, t),
                    ))
                    total_tx = db.query(models.Transaction).filter(
                        models.Transaction.workspace_id == transaction.workspace_id
                    ).count()
                    if total_tx == 1:
                        send_telegram_msg(chat_id, t('tip_multi'))
                    # Budget alert, streak, insight after confirm
                    try:
                        alert = _check_budget_alerts(transaction.workspace_id, category_id_for_msg, db, t)
                        if alert:
                            send_telegram_msg(chat_id, alert)
                        streak_msg = _check_streak(transaction.workspace_id, db, t)
                        if streak_msg:
                            send_telegram_msg(chat_id, streak_msg)
                        insight_msg = _generate_insight(transaction.workspace_id, category_id_for_msg, db, t)
                        if insight_msg:
                            send_telegram_msg(chat_id, insight_msg)
                        month_cmp = _check_month_comparison(transaction.workspace_id, db, t)
                        if month_cmp:
                            send_telegram_msg(chat_id, month_cmp)
                    except Exception as extra_err:
                        logger.warning("Erro extras post-confirm: %s", extra_err)
                
                logger.info("Callback de confirmacao processado com sucesso")
                return {'status': 'confirmed'}
                
            elif callback_data.startswith("cancel_"):
                # Cancelar transação
                pending_id_hex = callback_data.replace("cancel_", "")
                logger.info(f"Cancelando transação pendente: hex={pending_id_hex}, chat_id={chat_id}")
                
                # Buscar por hex curto (primeiros 16 caracteres do UUID)
                all_pending = db.query(models.TelegramPendingTransaction).filter(
                    models.TelegramPendingTransaction.chat_id == str(chat_id)
                ).all()
                
                logger.info(f"Transações pendentes encontradas para chat_id {chat_id}: {len(all_pending)}")
                
                pending = None
                for p in all_pending:
                    p_hex = p.id.hex[:16]
                    logger.info(f"Comparando: pending_id_hex={pending_id_hex}, p.id.hex[:16]={p_hex}, match={p.id.hex.startswith(pending_id_hex)}")
                    if p.id.hex.startswith(pending_id_hex):
                        pending = p
                        logger.info(f"Transação pendente encontrada: id={p.id}, description={p.description}, amount_cents={p.amount_cents}")
                        break
                
                if pending:
                    batch_id_hex = pending.batch_id.hex[:16] if pending.batch_id else None
                    db.delete(pending)
                    db.commit()
                    logger.info(f"Transação pendente eliminada com sucesso: id={pending.id}")
                    try:
                        requests.post(
                            f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/answerCallbackQuery",
                            json={'callback_query_id': callback_query['id']},
                            timeout=5,
                        )
                    except Exception as e:
                        logger.error(f"Erro ao responder callback query: {str(e)}")
                    if batch_id_hex:
                        message_id = callback_query.get("message", {}).get("message_id")
                        if message_id is not None:
                            msg_text, markup = _build_batch_message_and_keyboard(str(chat_id), batch_id_hex, db, t)
                            if msg_text and markup:
                                edit_telegram_message(chat_id, message_id, msg_text, markup)
                            else:
                                edit_telegram_message(chat_id, message_id, t('batch_list_empty'), {"inline_keyboard": []})
                    else:
                        send_telegram_msg(chat_id, t('transaction_cancelled'))
                    return {'status': 'cancelled'}
                else:
                    logger.warning(f"Transação pendente não encontrada: hex={pending_id_hex}, chat_id={chat_id}")
                    send_telegram_msg(chat_id, t('transaction_cancel_not_found'))
                    return {'status': 'not_found'}
            
            return {'status': 'ok'}
        
        # Processar mensagens normais
        if 'message' not in data:
            logger.info("Payload não contém 'message' - ignorando")
            return {'status': 'ignored'}
        
        logger.info("Processando mensagem normal")
        message = data['message']
        chat_id = message['chat']['id']
        text = message.get('text', '').strip()
        # Áudio/voice: transcrever com Whisper e usar o texto como mensagem
        if not text and (message.get('voice') or message.get('audio')):
            user_temp = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            language = (user_temp.language if user_temp and user_temp.language else None) or _telegram_lang(message.get("from"))
            t_audio = get_telegram_t(language)
            _send_typing_action(chat_id)
            try:
                text = transcribe_audio_from_telegram(message) or ""
            except OpenAIRateLimitError as e:
                logger.warning("[Telegram] Whisper rate limit (429): %s", e)
                send_telegram_msg(chat_id, t_audio('photo_rate_limit'))
                return {'status': 'error'}
            if not text:
                send_telegram_msg(chat_id, t_audio('audio_error'))
                return {'status': 'error'}
            text = text.strip()
            logger.info(f"Áudio transcrito: '{text[:100]}'")
        # Log claro: documento, foto ou texto
        if message.get('document'):
            doc = message['document']
            logger.info(f"Mensagem recebida: chat_id={chat_id}, document={doc.get('file_name', '')} (mime={doc.get('mime_type', '')})")
        elif message.get('photo'):
            logger.info(f"Mensagem recebida: chat_id={chat_id}, photo ({message['photo'][-1].get('file_size') or '?'} bytes)")
        else:
            logger.info(f"Mensagem recebida: chat_id={chat_id}, text='{text[:100] if text else ''}'")
        
        # Buscar utilizador para obter linguagem (se existir); senão usar idioma do Telegram
        user_temp = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
        language = (user_temp.language if user_temp and user_temp.language else None) or _telegram_lang(message.get("from"))
        t = get_telegram_t(language)
        
        # Verificar rate limit
        if not check_rate_limit(str(chat_id)):
            send_telegram_msg(chat_id, t('rate_limit'))
            return {'status': 'rate_limited'}
        
        # Comando /start
        if text.startswith('/start'):
            logger.info(f"Comando /start recebido de chat_id={chat_id}")
            user = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            logger.info(f"User encontrado: {user is not None}")
            
            if not user:
                # Primeira vez, pedir email (usar idioma do Telegram)
                lang_start = _telegram_lang(message.get("from"))
                t_start = get_telegram_t(lang_start)
                send_telegram_msg(chat_id, t_start('welcome_new'))
                return {'status': 'email_required'}
            else:
                # Já associado - enviar mensagem de boas-vindas e fixar
                language = user.language if user.language else 'pt'
                t_start = get_telegram_t(language)
                send_telegram_msg(chat_id, t_start('welcome_return'), pin_message=True)
                return {'status': 'ok'}
        
        # Comandos /info, /help e /ajuda
        if text.startswith('/info') or text.startswith('/help') or text.startswith('/ajuda'):
            send_telegram_msg(chat_id, t('help_guide_v2') or t('help_guide'))
            return {'status': 'ok'}
        
        # Comando /resumo ou /hoje - Resumo do dia
        if text.strip().lower() in ('/resumo', '/hoje'):
            user = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            workspace = db.query(models.Workspace).filter(models.Workspace.owner_id == user.id).first()
            if not workspace:
                send_telegram_msg(chat_id, t('workspace_not_found'))
                return {'status': 'error'}
            today = date.today()
            q = db.query(
                func.sum(case((models.Transaction.amount_cents < 0, models.Transaction.amount_cents), else_=0)),
                func.sum(case((models.Transaction.amount_cents > 0, models.Transaction.amount_cents), else_=0)),
                func.count(models.Transaction.id),
            ).filter(
                models.Transaction.workspace_id == workspace.id,
                models.Transaction.transaction_date == today,
            ).first()
            expenses_cents = int(q[0] or 0)
            income_cents = int(q[1] or 0)
            count = int(q[2] or 0)
            expenses = abs(expenses_cents) / 100
            income = income_cents / 100
            balance = income - abs(expenses_cents) / 100
            # Garantir sempre 2 casas decimais na apresentação
            expenses_str = f"{expenses:.2f}"
            income_str = f"{income:.2f}"
            balance_str = f"{balance:.2f}"
            lang = user.language if user.language else 'pt'
            t_sum = get_telegram_t(lang)
            if count == 0:
                send_telegram_msg(chat_id, t_sum('summary_empty'))
            else:
                send_telegram_msg(
                    chat_id,
                    t_sum('summary_today').format(
                        expenses=expenses_str,
                        income=income_str,
                        count=count,
                        balance=balance_str,
                    ),
                )
            return {'status': 'ok'}
        
        # Comando /mes - Resumo do mês
        if text.strip().lower() == '/mes':
            user = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            workspace = db.query(models.Workspace).filter(models.Workspace.owner_id == user.id).first()
            if not workspace:
                send_telegram_msg(chat_id, t('workspace_not_found'))
                return {'status': 'error'}
            today = date.today()
            first_day = today.replace(day=1)
            q = db.query(
                func.sum(case((models.Transaction.amount_cents < 0, models.Transaction.amount_cents), else_=0)),
                func.sum(case((models.Transaction.amount_cents > 0, models.Transaction.amount_cents), else_=0)),
                func.count(models.Transaction.id),
            ).filter(
                models.Transaction.workspace_id == workspace.id,
                models.Transaction.transaction_date >= first_day,
                models.Transaction.transaction_date <= today,
            ).first()
            expenses_cents = int(q[0] or 0)
            income_cents = int(q[1] or 0)
            count = int(q[2] or 0)
            expenses = abs(expenses_cents) / 100
            income = income_cents / 100
            balance = income - abs(expenses_cents) / 100
            # Garantir sempre 2 casas decimais na apresentação
            expenses_str = f"{expenses:.2f}"
            income_str = f"{income:.2f}"
            balance_str = f"{balance:.2f}"
            lang = user.language if user.language else 'pt'
            t_sum = get_telegram_t(lang)
            if count == 0:
                send_telegram_msg(chat_id, t_sum('summary_empty'))
            else:
                send_telegram_msg(
                    chat_id,
                    t_sum('summary_month').format(
                        expenses=expenses_str,
                        income=income_str,
                        count=count,
                        balance=balance_str,
                    ),
                )
            return {'status': 'ok'}
        
        # Comando /semana - Resumo da semana
        if text.strip().lower() == '/semana':
            user = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            workspace = db.query(models.Workspace).filter(models.Workspace.owner_id == user.id).first()
            if not workspace:
                send_telegram_msg(chat_id, t('workspace_not_found'))
                return {'status': 'error'}
            lang = user.language if user.language else 'pt'
            t_sum = get_telegram_t(lang)
            today = date.today()
            week_start = today - timedelta(days=today.weekday())
            q = db.query(
                func.sum(case((models.Transaction.amount_cents < 0, models.Transaction.amount_cents), else_=0)),
                func.sum(case((models.Transaction.amount_cents > 0, models.Transaction.amount_cents), else_=0)),
                func.count(models.Transaction.id),
            ).filter(
                models.Transaction.workspace_id == workspace.id,
                models.Transaction.transaction_date >= week_start,
                models.Transaction.transaction_date <= today,
            ).first()
            expenses_cents = int(q[0] or 0)
            income_cents = int(q[1] or 0)
            count = int(q[2] or 0)
            if count == 0:
                send_telegram_msg(chat_id, t_sum('summary_empty'))
                return {'status': 'ok'}
            expenses = abs(expenses_cents) / 100
            income = income_cents / 100
            balance = income - expenses
            # Top category this week
            top_cat = db.query(
                models.Category.name,
                func.sum(func.abs(models.Transaction.amount_cents)).label('total'),
            ).join(models.Category, models.Transaction.category_id == models.Category.id).filter(
                models.Transaction.workspace_id == workspace.id,
                models.Transaction.transaction_date >= week_start,
                models.Transaction.transaction_date <= today,
                models.Transaction.amount_cents < 0,
            ).group_by(models.Category.name).order_by(func.sum(func.abs(models.Transaction.amount_cents)).desc()).first()
            top_category = top_cat[0] if top_cat else "-"
            top_amount = f"{(top_cat[1] or 0) / 100:.2f}" if top_cat else "0.00"
            # Previous week comparison
            prev_week_start = week_start - timedelta(days=7)
            prev_week_end = week_start - timedelta(days=1)
            prev_q = db.query(
                func.sum(case((models.Transaction.amount_cents < 0, models.Transaction.amount_cents), else_=0)),
            ).filter(
                models.Transaction.workspace_id == workspace.id,
                models.Transaction.transaction_date >= prev_week_start,
                models.Transaction.transaction_date <= prev_week_end,
            ).scalar()
            prev_expenses = abs(int(prev_q or 0)) / 100
            comparison = ""
            if prev_expenses > 0:
                diff = abs(expenses - prev_expenses)
                if expenses < prev_expenses:
                    comparison = t_sum('week_comparison_better').format(diff=f"{diff:.2f}")
                elif expenses > prev_expenses:
                    comparison = t_sum('week_comparison_worse').format(diff=f"{diff:.2f}")
            send_telegram_msg(chat_id, t_sum('summary_week').format(
                expenses=f"{expenses:.2f}",
                income=f"{income:.2f}",
                count=count,
                top_category=top_category,
                top_amount=top_amount,
                comparison=comparison,
                balance=f"{balance:.2f}",
            ))
            return {'status': 'ok'}
        
        # Comando /pendentes - Listar transações pendentes
        if text.strip().lower() == '/pendentes':
            user = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            workspace = db.query(models.Workspace).filter(models.Workspace.owner_id == user.id).first()
            if not workspace:
                send_telegram_msg(chat_id, t('workspace_not_found'))
                return {'status': 'error'}
            lang = user.language if user.language else 'pt'
            t_pend = get_telegram_t(lang)
            pendents = db.query(models.TelegramPendingTransaction).filter(
                models.TelegramPendingTransaction.chat_id == str(chat_id),
                models.TelegramPendingTransaction.workspace_id == workspace.id,
            ).order_by(models.TelegramPendingTransaction.created_at).all()
            if not pendents:
                send_telegram_msg(chat_id, t_pend('pendentes_empty'))
                return {'status': 'ok'}
            lines = []
            for p in pendents:
                cat = db.query(models.Category).filter(models.Category.id == p.category_id).first()
                cat_name = cat.name if cat else "Outros"
                desc_display = _html_escape(_shorten_description_for_list(p.description))
                amount_display = "{:.2f}".format(abs(p.amount_cents) / 100).replace(".", ",")
                lines.append(t_pend('list_pending_line').format(description=desc_display, amount=amount_display, category=cat_name))
            send_telegram_msg(chat_id, t_pend('pendentes_list').format(count=len(pendents), lines="".join(lines)))
            return {'status': 'ok'}
        
        # Comando /revoke - Desvincular Telegram da conta
        if text.strip().lower() == '/revoke':
            user = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            lang = user.language if user.language else 'pt'
            t_revoke = get_telegram_t(lang)
            logger.info(f"[AÇÃO SENSÍVEL] /revoke: chat_id={chat_id}, user_id={user.id}, email={getattr(user, 'email', '')[:10]}***")
            user.phone_number = None
            db.query(models.TelegramPendingTransaction).filter(
                models.TelegramPendingTransaction.chat_id == str(chat_id),
            ).delete(synchronize_session=False)
            db.commit()
            send_telegram_msg(chat_id, t_revoke('revoke_ok'))
            return {'status': 'ok'}
        
        # Comando /idioma pt|en - Mudar idioma
        if text.strip().lower().startswith('/idioma '):
            user = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            part = text.strip().split(maxsplit=1)
            lang_arg = (part[1].strip().lower() if len(part) > 1 else "") or "pt"
            if lang_arg in ("en", "english"):
                user.language = "en"
                db.commit()
                t_lang = get_telegram_t("en")
                send_telegram_msg(chat_id, t_lang('language_set_en'))
            else:
                user.language = "pt"
                db.commit()
                t_lang = get_telegram_t("pt")
                send_telegram_msg(chat_id, t_lang('language_set'))
            return {'status': 'ok'}
        
        # Comando /categoria [nome] ou /definir [nome] - Categoria por defeito até "stop"
        if text.strip().lower().startswith('/categoria ') or text.strip().lower().startswith('/definir '):
            user_cat = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user_cat:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            workspace_cat = db.query(models.Workspace).filter(models.Workspace.owner_id == user_cat.id).first()
            if not workspace_cat:
                send_telegram_msg(chat_id, t('workspace_not_found'))
                return {'status': 'error'}
            part = text.strip().split(maxsplit=1)
            rest = (part[1].strip() if len(part) > 1 else "") or ""
            lang_cat = user_cat.language if user_cat.language else 'pt'
            t_cat = get_telegram_t(lang_cat)
            if rest.lower() in ("stop", "parar", "clear", "nenhuma", "none", "remove"):
                user_cat.telegram_default_category_id = None
                db.commit()
                send_telegram_msg(chat_id, t_cat('categoria_default_cleared'))
                return {'status': 'ok'}
            cat_by_name = db.query(models.Category).filter(
                models.Category.workspace_id == workspace_cat.id,
                func.lower(models.Category.name) == rest.lower(),
            ).first()
            if not cat_by_name:
                send_telegram_msg(chat_id, t_cat('categoria_not_found').format(name=rest[:50]))
                return {'status': 'not_found'}
            user_cat.telegram_default_category_id = cat_by_name.id
            db.commit()
            send_telegram_msg(chat_id, t_cat('categoria_default_set').format(name=cat_by_name.name))
            return {'status': 'ok'}
        
        # Comando /clear - Limpar transações pendentes
        if text.startswith('/clear'):
            logger.info(f"Comando /clear recebido de chat_id={chat_id}")
            user = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            
            language = user.language if user.language else 'pt'
            t_clear = get_telegram_t(language)
            
            workspace = db.query(models.Workspace).filter(models.Workspace.owner_id == user.id).first()
            if not workspace:
                send_telegram_msg(chat_id, t_clear('workspace_not_found'))
                return {'status': 'error'}
            
            # Eliminar todas as transações pendentes do utilizador
            pending_transactions = db.query(models.TelegramPendingTransaction).filter(
                models.TelegramPendingTransaction.chat_id == str(chat_id),
                models.TelegramPendingTransaction.workspace_id == workspace.id
            ).all()
            
            count = len(pending_transactions)
            if count > 0:
                for pending in pending_transactions:
                    db.delete(pending)
                db.commit()
                logger.info(f"[AÇÃO SENSÍVEL] /clear em massa: chat_id={chat_id}, user_id={user.id}, count={count}")
                send_telegram_msg(chat_id, t_clear('clear_success').format(count=count))
            else:
                send_telegram_msg(chat_id, t_clear('clear_empty'))
            
            return {'status': 'ok'}
        
        # Comando /exportar - Exportar transações em CSV
        if text.strip().lower().startswith('/exportar'):
            user_exp = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user_exp:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            workspace_exp = db.query(models.Workspace).filter(models.Workspace.owner_id == user_exp.id).first()
            if not workspace_exp:
                send_telegram_msg(chat_id, t('workspace_not_found'))
                return {'status': 'error'}
            lang_exp = user_exp.language if user_exp.language else 'pt'
            t_exp = get_telegram_t(lang_exp)
            parts = text.strip().lower().split()
            period = parts[1] if len(parts) > 1 else None
            today = date.today()
            if period in ('semana', 'week'):
                start_date = today - timedelta(days=today.weekday())
                period_label = "semana" if lang_exp == 'pt' else "week"
            elif period in ('mes', 'month', 'mês'):
                start_date = today.replace(day=1)
                period_label = "mês" if lang_exp == 'pt' else "month"
            else:
                send_telegram_msg(chat_id, t_exp('export_usage'))
                return {'status': 'ok'}
            transactions = db.query(models.Transaction).join(
                models.Category, models.Transaction.category_id == models.Category.id
            ).filter(
                models.Transaction.workspace_id == workspace_exp.id,
                models.Transaction.transaction_date >= start_date,
                models.Transaction.transaction_date <= today,
            ).order_by(models.Transaction.transaction_date.desc()).all()
            if not transactions:
                send_telegram_msg(chat_id, t_exp('export_no_data'))
                return {'status': 'ok'}
            csv_lines = ["Data,Descrição,Valor,Categoria,Tipo" if lang_exp == 'pt' else "Date,Description,Amount,Category,Type"]
            for tx in transactions:
                cat = db.query(models.Category).filter(models.Category.id == tx.category_id).first()
                cat_name = cat.name if cat else "-"
                tx_type = "Despesa" if tx.amount_cents < 0 else "Receita"
                if lang_exp == 'en':
                    tx_type = "Expense" if tx.amount_cents < 0 else "Income"
                desc = (tx.description or "").replace(",", ";").replace('"', "'")
                csv_lines.append(f"{tx.transaction_date},{desc},{abs(tx.amount_cents)/100:.2f},{cat_name},{tx_type}")
            csv_content = "\n".join(csv_lines)
            try:
                csv_bytes = csv_content.encode('utf-8-sig')
                files = {'document': (f"finly_{period_label}_{today.isoformat()}.csv", io.BytesIO(csv_bytes), 'text/csv')}
                url_doc = f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/sendDocument"
                requests.post(url_doc, data={"chat_id": chat_id}, files=files, timeout=15)
            except Exception as e:
                logger.error(f"Erro ao enviar CSV: {e}")
                send_telegram_msg(chat_id, t_exp('export_error'))
            return {'status': 'ok'}
        
        # Comando /recorrentes - Listar despesas recorrentes
        if text.strip().lower() == '/recorrentes':
            user_rec = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user_rec:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            workspace_rec = db.query(models.Workspace).filter(models.Workspace.owner_id == user_rec.id).first()
            if not workspace_rec:
                send_telegram_msg(chat_id, t('workspace_not_found'))
                return {'status': 'error'}
            lang_rec = user_rec.language if user_rec.language else 'pt'
            t_rec = get_telegram_t(lang_rec)
            recurrings = db.query(models.RecurringTransaction).filter(
                models.RecurringTransaction.workspace_id == workspace_rec.id,
                models.RecurringTransaction.is_active == True,
            ).all()
            if not recurrings:
                send_telegram_msg(chat_id, t_rec('recurring_list_empty'))
                return {'status': 'ok'}
            msg = t_rec('recurring_list_header')
            total = 0
            for r in recurrings:
                amount = abs(r.amount_cents) / 100
                total += amount
                msg += t_rec('recurring_list_line').format(
                    description=r.description or "-",
                    amount=f"{amount:.2f}",
                    day=r.day_of_month or 1,
                )
            msg += f"\n💰 <b>Total:</b> {total:.2f}€/mês"
            send_telegram_msg(chat_id, msg)
            return {'status': 'ok'}
        
        # Comando /recorrente - Criar despesa recorrente
        if text.strip().lower().startswith('/recorrente '):
            user_rec = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user_rec:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            workspace_rec = db.query(models.Workspace).filter(models.Workspace.owner_id == user_rec.id).first()
            if not workspace_rec:
                send_telegram_msg(chat_id, t('workspace_not_found'))
                return {'status': 'error'}
            lang_rec = user_rec.language if user_rec.language else 'pt'
            t_rec = get_telegram_t(lang_rec)
            rest_text = text.strip().split(maxsplit=1)[1] if len(text.strip().split(maxsplit=1)) > 1 else ""
            amount_match = re.search(r'(\d+[.,]?\d*)\s*€?', rest_text)
            day_match = re.search(r'dia\s*(\d{1,2})|day\s*(\d{1,2})', rest_text, re.IGNORECASE)
            if not amount_match:
                send_telegram_msg(chat_id, t_rec('recurring_parse_error'))
                return {'status': 'error'}
            amount_val = float(amount_match.group(1).replace(',', '.'))
            day_of_month = int(day_match.group(1) or day_match.group(2)) if day_match else 1
            day_of_month = max(1, min(28, day_of_month))
            desc_text = rest_text[:amount_match.start()].strip()
            if not desc_text:
                desc_text = rest_text[amount_match.end():].strip()
                if day_match:
                    desc_text = desc_text[:day_match.start()].strip() if day_match.start() > 0 else desc_text[day_match.end():].strip()
            desc_text = re.sub(r'\s*dia\s*\d{1,2}\s*|\s*day\s*\d{1,2}\s*', '', desc_text, flags=re.IGNORECASE).strip()
            if not desc_text:
                desc_text = "Recorrente"
            # Find first expense category
            default_cat = db.query(models.Category).filter(
                models.Category.workspace_id == workspace_rec.id,
                models.Category.type == 'expense',
            ).first()
            new_recurring = models.RecurringTransaction(
                workspace_id=workspace_rec.id,
                category_id=default_cat.id if default_cat else None,
                description=desc_text[:100],
                amount_cents=int(-abs(amount_val) * 100),
                day_of_month=day_of_month,
                is_active=True,
            )
            db.add(new_recurring)
            db.commit()
            send_telegram_msg(chat_id, t_rec('recurring_created').format(
                description=desc_text[:100],
                amount=f"{amount_val:.2f}",
                day=day_of_month,
            ))
            return {'status': 'ok'}
        
        # Comando /desfazer - Anular última transação (últimos 5 min)
        if text.strip().lower() in ('/desfazer', '/undo'):
            user_undo = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
            if not user_undo:
                send_telegram_msg(chat_id, t('clear_unauthorized'))
                return {'status': 'unauthorized'}
            workspace_undo = db.query(models.Workspace).filter(models.Workspace.owner_id == user_undo.id).first()
            if not workspace_undo:
                send_telegram_msg(chat_id, t('workspace_not_found'))
                return {'status': 'error'}
            lang_undo = user_undo.language if user_undo.language else 'pt'
            t_undo = get_telegram_t(lang_undo)
            five_min_ago = datetime.now(timezone.utc) - timedelta(minutes=5)
            last_tx = db.query(models.Transaction).filter(
                models.Transaction.workspace_id == workspace_undo.id,
            ).order_by(models.Transaction.created_at.desc()).first()
            if not last_tx:
                send_telegram_msg(chat_id, t_undo('undo_no_recent'))
                return {'status': 'ok'}
            if last_tx.created_at and last_tx.created_at.replace(tzinfo=timezone.utc if last_tx.created_at.tzinfo is None else last_tx.created_at.tzinfo) < five_min_ago:
                send_telegram_msg(chat_id, t_undo('undo_expired'))
                return {'status': 'ok'}
            desc = last_tx.description or "-"
            amount = abs(last_tx.amount_cents) / 100
            db.delete(last_tx)
            db.commit()
            send_telegram_msg(chat_id, t_undo('undo_success').format(
                description=desc,
                amount=f"{amount:.2f}",
            ))
            return {'status': 'ok'}
        
        # Processar email (associação)
        if "@" in text and "." in text:
            logger.info(f"Email detectado na mensagem: {text[:50]}")
            email_limpo = text.lower().replace(" ", "").strip()
            logger.info(f"Email limpo: {email_limpo[:10]}***")
            
            # Validar formato (usar idioma do Telegram antes de encontrar user)
            lang_email = _telegram_lang(message.get("from"))
            t_email = get_telegram_t(lang_email)
            if not validate_email(email_limpo):
                logger.warning(f"Email inválido: {email_limpo}")
                send_telegram_msg(chat_id, t_email('invalid_email'))
                return {'status': 'invalid_email'}
            
            # Procurar utilizador (insensível a maiúsculas: mesma conta que no PC)
            user = db.query(models.User).filter(func.lower(models.User.email) == email_limpo).first()
            
            if not user:
                # Resposta genérica para prevenir email enumeration
                send_telegram_msg(chat_id, t_email('email_not_found'))
                logger.warning(f"Tentativa de associação com email não registado: {email_limpo[:5]}***")
                return {'status': 'not_found'}
            
            # Obter linguagem do utilizador encontrado
            language = user.language if user.language else 'pt'
            t_email = get_telegram_t(language)
            
            # Verificar se é conta Pro (admin, subscrição ativa ou Pro concedido até data)
            if not user.has_effective_pro():
                send_telegram_msg(chat_id, t_email('pro_required'))
                return {'status': 'pro_required'}
            
            # Verificar conflitos (um chat_id só pode estar associado a um email)
            existing_user = db.query(models.User).filter(
                models.User.phone_number == str(chat_id)
            ).first()
            
            if existing_user and (existing_user.email or "").lower() != email_limpo:
                # Já está associado a outro email
                send_telegram_msg(chat_id, t_email('already_associated').format(email=f"{existing_user.email[:3]}***"))
                return {'status': 'already_associated'}
            
            # Associar Telegram (armazenar chat_id em phone_number)
            old_phone = user.phone_number
            user.phone_number = str(chat_id)
            db.commit()
            
            # Verificar workspace após associação
            workspace_check = db.query(models.Workspace).filter(models.Workspace.owner_id == user.id).first()
            logger.info(f"Conta Telegram associada: email={email_limpo[:10]}***, user_id={user.id}, workspace_id={workspace_check.id if workspace_check else None}, chat_id={chat_id}")
            
            send_telegram_msg(chat_id, t_email('account_linked_success').format(email=f"{user.email[:3]}***"), pin_message=True)
            # Mini-tutorial após linkagem
            send_telegram_msg(chat_id, t_email('welcome_tutorial'))
            return {'status': 'ok'}
        
        # Procurar User
        logger.info(f"Buscando user com phone_number={chat_id}")
        user = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
        logger.info(f"User encontrado: {user is not None} (id: {user.id if user else None}, email: {user.email[:10] if user else None}***)")
        if user:
            logger.info(f"telegram_auto_confirm: {user.telegram_auto_confirm}")
            # Verificar workspace do user
            workspace_check = db.query(models.Workspace).filter(models.Workspace.owner_id == user.id).first()
            logger.info(f"Workspace do user Telegram: {workspace_check.id if workspace_check else None}")
            # Atualizar t com a linguagem do utilizador
            language = user.language if user.language else 'pt'
            t = get_telegram_t(language)
        if not user:
            send_telegram_msg(chat_id, t('unauthorized'))
            return {'status': 'unauthorized'}
        
        logger.info(f"Buscando workspace para user_id={user.id}")
        workspace = db.query(models.Workspace).filter(models.Workspace.owner_id == user.id).first()
        logger.info(f"Workspace encontrado: {workspace is not None} (id: {workspace.id if workspace else None})")
        if not workspace:
            send_telegram_msg(chat_id, t('workspace_not_found'))
            return {'status': 'error'}
        
        parsed = None
        # Processar documento: imagem (Vision), PDF, CSV ou Excel
        if 'document' in message:
            doc = message['document']
            file_id = doc.get('file_id')
            file_name = (doc.get('file_name') or '').lower()
            mime = (doc.get('mime_type') or '').lower()
            is_image = mime.startswith('image/') or file_name.endswith(('.jpg', '.jpeg', '.png', '.webp'))
            is_pdf = mime == 'application/pdf' or file_name.endswith('.pdf')
            is_csv = mime == 'text/csv' or file_name.endswith('.csv') or 'csv' in mime
            is_xlsx = (
                mime in ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel')
                or file_name.endswith('.xlsx') or file_name.endswith('.xls')
            )

            if is_image:
                _send_typing_action(chat_id)
                all_categories = db.query(models.Category).filter(models.Category.workspace_id == workspace.id).all()
                try:
                    photo_result = process_photo_with_openai(file_id, all_categories)
                except OpenAIRateLimitError as e:
                    logger.warning("[Telegram] OpenAI rate limit (document image): %s", e)
                    send_telegram_msg(chat_id, t('photo_rate_limit'))
                    return {'status': 'error'}
                if not photo_result:
                    send_telegram_msg(chat_id, t('photo_not_supported'))
                    return {'status': 'error'}
                try:
                    parsed = _build_parsed_from_photo_result(photo_result, workspace, db)
                except AIUnavailableError as ae:
                    logger.warning("[Telegram] Vision AIUnavailableError: %s", ae)
                    send_telegram_msg(chat_id, t('ai_unavailable'))
                    return {'status': 'error'}
                if not parsed:
                    send_telegram_msg(chat_id, t('document_no_data'))
                    return {'status': 'error'}
                if parsed.get("multiple"):
                    logger.info("[Telegram] Documento imagem processado: %s transações", len(parsed.get("transactions", [])))
                else:
                    logger.info("[Telegram] Documento imagem processado: 1 transação - %s", parsed.get("description"))

            elif is_pdf or is_csv or is_xlsx:
                _send_typing_action(chat_id)
                content = _download_telegram_file(file_id)
                if not content:
                    send_telegram_msg(chat_id, t('document_not_supported'))
                    return {'status': 'error'}
                if is_pdf:
                    extracted = _extract_text_from_pdf(content)
                elif is_csv:
                    extracted = _extract_text_from_csv(content)
                else:
                    extracted = _extract_text_from_xlsx(content)
                if not extracted or len(extracted.strip()) < 5:
                    send_telegram_msg(chat_id, t('document_no_data'))
                    return {'status': 'error'}
                default_cat_id = getattr(user, 'telegram_default_category_id', None)
                extracted_stripped = extracted.strip()
                use_ai_for_document = len(extracted_stripped) >= 400 and getattr(settings, "OPENAI_API_KEY", None)
                parsed = None
                if use_ai_for_document:
                    try:
                        ai_doc_result = parse_document_with_openai(extracted_stripped, workspace, db, default_cat_id)
                        if ai_doc_result:
                            parsed = _build_parsed_from_photo_result(ai_doc_result, workspace, db)
                            if parsed:
                                logger.info("[Telegram] Documento processado com IA: %s transações", len(parsed.get("transactions", [parsed])))
                    except OpenAIRateLimitError:
                        send_telegram_msg(chat_id, t('ai_busy'))
                        return {'status': 'error'}
                    except Exception as ai_err:
                        logger.warning("[Telegram] IA para documento falhou, fallback regex: %s", ai_err)
                if not parsed:
                    try:
                        parsed = parse_transaction(extracted_stripped, workspace, db, default_category_id=default_cat_id)
                    except AIUnavailableError as ae:
                        err_str = str(ae).lower()
                        if "429" in err_str or "rate" in err_str or "limit" in err_str:
                            send_telegram_msg(chat_id, t('ai_busy'))
                        else:
                            send_telegram_msg(chat_id, t('ai_unavailable'))
                        return {'status': 'error'}
                if not parsed:
                    send_telegram_msg(chat_id, t('document_no_data'))
                    return {'status': 'error'}
                if parsed.get("multiple"):
                    logger.info("[Telegram] Documento %s processado: %s transações", "PDF" if is_pdf else "CSV/Excel", len(parsed.get("transactions", [])))
                else:
                    logger.info("[Telegram] Documento processado: 1 transação - %s", parsed.get("description"))

            else:
                send_telegram_msg(chat_id, t('document_not_supported'))
                return {'status': 'error'}

        # Processar fotos (OpenAI Vision)
        elif 'photo' in message:
            file_id = message['photo'][-1]['file_id']
            _send_typing_action(chat_id)
            # Carregar todas as categorias do workspace para enviar no prompt da Vision
            all_categories = db.query(models.Category).filter(
                models.Category.workspace_id == workspace.id
            ).all()
            logger.info("[Telegram] Foto recebida file_id=%s workspace_id=%s categorias=%s", file_id[:20] if file_id else None, workspace.id, len(all_categories))
            try:
                photo_result = process_photo_with_openai(file_id, all_categories)
            except OpenAIRateLimitError as e:
                logger.warning("[Telegram] OpenAI rate limit (429): %s -> enviando photo_rate_limit", e)
                send_telegram_msg(chat_id, t('photo_rate_limit'))
                return {'status': 'error'}
            if not photo_result:
                logger.warning("[Telegram] process_photo_with_openai retornou None -> enviando photo_not_supported")
                send_telegram_msg(chat_id, t('photo_not_supported'))
                return {'status': 'error'}
            logger.info("[Telegram] Vision OK, a construir parsed (pode ser lista)")
            try:
                parsed = _build_parsed_from_photo_result(photo_result, workspace, db)
            except AIUnavailableError as ae:
                logger.warning("[Telegram] Vision AIUnavailableError: %s", ae)
                send_telegram_msg(chat_id, t('ai_unavailable'))
                return {'status': 'error'}
            if not parsed:
                logger.warning("[Telegram] Vision retornou None -> enviando photo_not_supported")
                send_telegram_msg(chat_id, t('photo_not_supported'))
                return {'status': 'error'}
            if parsed.get("multiple"):
                logger.info("[Telegram] Foto processada: %s transações (extrato/tabela)", len(parsed.get("transactions", [])))
            else:
                logger.info("[Telegram] Foto processada: 1 transação - %s", parsed.get("description"))
        # Processar texto: AI Router (GPT) ou fast-path regex
        elif text:
            text = correct_transcription_with_history(text, workspace.id, db)
            logger.info(f"Processando texto: '{text}'")
            default_cat_id = getattr(user, 'telegram_default_category_id', None)

            # Emoji detection: se o texto contém emoji mapeado, usar como categoria
            if not default_cat_id:
                all_cats = db.query(models.Category).filter(models.Category.workspace_id == workspace.id).all()
                emoji_cat, emoji_found = _match_emoji_to_category(text, all_cats, user.language or 'pt')
                if emoji_cat:
                    default_cat_id = emoji_cat.id
                    logger.info("[Emoji] Emoji %s -> categoria %s", emoji_found, emoji_cat.name)

            # FAST-PATH: mensagens com valor monetario obvio -> parse_transaction direto (sem GPT)
            if _is_obvious_transaction(text):
                logger.info("[Fast-path] Mensagem com valor obvio, a usar parse_transaction direto")
                try:
                    parsed = parse_transaction(text, workspace, db, default_category_id=default_cat_id)
                except AIUnavailableError as ae:
                    err_str = str(ae).lower()
                    if "429" in err_str or "rate" in err_str or "limit" in err_str:
                        send_telegram_msg(chat_id, t('ai_busy'))
                    else:
                        send_telegram_msg(chat_id, t('ai_unavailable'))
                    return {'status': 'error'}

                if parsed:
                    logger.info(f"[Fast-path] Resultado do parsing: {parsed}")
                else:
                    logger.warning(f"[Fast-path] parse_transaction falhou, a tentar AI Router")
                    # Fallthrough: se o regex falha mesmo com valor, tentar o AI Router

            # AI ROUTER: GPT-4o-mini decide intent (transacao, pergunta, conselho, chat)
            if not parsed:
                if not _check_gpt_rate_limit(str(chat_id)):
                    logger.warning("[AI Router] Rate limit GPT para chat_id=%s", chat_id)
                    send_telegram_msg(chat_id, t('ai_rate_limited'))
                    return {'status': 'gpt_rate_limited'}

                # Enviar indicador "a escrever..." (três pontos nativos do Telegram)
                _send_typing_action(chat_id)

                # Guardar mensagem do user na memória de conversa
                _add_to_memory(str(chat_id), "user", text)

                result = await ai_route_message(text, str(chat_id), user, workspace, db, t)
                intent = result.get("intent", "fallback")
                logger.info("[AI Router] intent=%s para texto='%s'", intent, text[:60])

                if intent == "gpt_rate_limited":
                    send_telegram_msg(chat_id, t('ai_rate_limited'))
                    return {'status': 'gpt_rate_limited'}

                if intent == "transaction":
                    # GPT extraiu transacoes -> processar
                    ai_result = _handle_ai_transaction(
                        result.get("transactions", []),
                        str(chat_id), user, workspace, db, t,
                        original_text=text,
                    )
                    if ai_result["status"] == "no_valid_transactions":
                        # GPT pensou que era transacao mas nao conseguiu extrair -> fallback regex
                        try:
                            parsed = parse_transaction(text, workspace, db, default_category_id=default_cat_id)
                        except AIUnavailableError:
                            pass
                        if not parsed:
                            send_telegram_msg(chat_id, t('ai_error'))
                            return {'status': 'error'}
                    elif ai_result["status"] == "multiple":
                        parsed = ai_result["parsed"]
                    elif ai_result["status"] == "single":
                        parsed = ai_result["parsed"]
                    else:
                        send_telegram_msg(chat_id, t('ai_error'))
                        return {'status': 'error'}

                elif intent in ("question", "advice", "chat"):
                    response_text = result.get("response", "")
                    if response_text:
                        send_telegram_msg(chat_id, response_text)
                        _add_to_memory(str(chat_id), "assistant", response_text)
                    return {'status': 'success'}

                elif intent == "fallback":
                    # GPT falhou -> fallback para parse_transaction
                    logger.info("[AI Router] Fallback para parse_transaction")
                    try:
                        parsed = parse_transaction(text, workspace, db, default_category_id=default_cat_id)
                    except AIUnavailableError as ae:
                        err_str = str(ae).lower()
                        if "429" in err_str or "rate" in err_str or "limit" in err_str:
                            send_telegram_msg(chat_id, t('ai_busy'))
                        else:
                            send_telegram_msg(chat_id, t('ai_unavailable'))
                        return {'status': 'error'}

            if not parsed:
                logger.warning(f"Nao foi possivel fazer parse da mensagem: '{text}'")
                send_telegram_msg(chat_id, t('parse_error'))
                return {'status': 'error'}
        
        if not parsed:
            logger.info("Mensagem não processada (sem texto nem foto)")
            return {'status': 'ignored'}
        
        # Limite de pendentes: máx 20; acima disso pedir para confirmar ou /clear
        if not user.telegram_auto_confirm:
            pending_count = db.query(models.TelegramPendingTransaction).filter(
                models.TelegramPendingTransaction.chat_id == str(chat_id),
                models.TelegramPendingTransaction.workspace_id == workspace.id,
            ).count()
            if pending_count >= 20:
                send_telegram_msg(chat_id, t('too_many_pending').format(count=pending_count))
                return {'status': 'too_many_pending'}
            # Aviso de pendentes antigos (>24h)
            stale_cutoff = datetime.now(timezone.utc) - timedelta(hours=PENDING_STALE_HOURS)
            stale_count = db.query(models.TelegramPendingTransaction).filter(
                models.TelegramPendingTransaction.chat_id == str(chat_id),
                models.TelegramPendingTransaction.workspace_id == workspace.id,
                models.TelegramPendingTransaction.created_at < stale_cutoff,
            ).count()
            if stale_count > 0:
                send_telegram_msg(chat_id, t('pending_stale'))
        
        # Processar múltiplas transações (lista): uma mensagem com todas as linhas + total + Confirmar tudo / Cancelar tudo
        if parsed.get('multiple'):
            transactions = parsed['transactions']
            created_count = 0
            batch_id = uuid.uuid4()
            batch_id_hex = batch_id.hex[:16]
            # Evitar duplicados (ex.: PDF com mesma lista duas vezes ou parser a repetir linhas)
            # Incluir descrição na chave para não descartar transações diferentes com mesmo valor e data (ex.: duas compras 20€ no mesmo dia)
            seen_key = set()

            for trans_data in transactions:
                amount_cents = int(trans_data['amount'] * 100)
                if trans_data.get('is_vault', False):
                    if trans_data.get('is_vault_withdrawal', False):
                        amount_cents = -abs(amount_cents)
                    else:
                        amount_cents = abs(amount_cents)
                elif trans_data['type'] == 'expense':
                    amount_cents = -abs(amount_cents)
                else:
                    amount_cents = abs(amount_cents)

                trans_date = trans_data.get('transaction_date') or date.today()
                desc = (trans_data.get('description') or '').strip()[:200]
                dedup_key = (amount_cents, trans_date, desc)
                if dedup_key in seen_key:
                    continue
                seen_key.add(dedup_key)

                if user.telegram_auto_confirm:
                    transaction = models.Transaction(
                        workspace_id=workspace.id,
                        category_id=trans_data['category_id'],
                        amount_cents=amount_cents,
                        description=trans_data['description'],
                        inference_source=trans_data.get('inference_source'),
                        decision_reason=trans_data.get('decision_reason'),
                        needs_review=trans_data.get('needs_review', False),
                        transaction_date=trans_date,
                    )
                    db.add(transaction)
                    created_count += 1
                else:
                    pending = models.TelegramPendingTransaction(
                        chat_id=str(chat_id),
                        workspace_id=workspace.id,
                        category_id=trans_data['category_id'],
                        amount_cents=amount_cents,
                        description=trans_data['description'],
                        inference_source=trans_data.get('inference_source'),
                        decision_reason=trans_data.get('decision_reason'),
                        needs_review=trans_data.get('needs_review', False),
                        transaction_date=trans_date,
                        batch_id=batch_id,
                    )
                    db.add(pending)

            if user.telegram_auto_confirm:
                db.commit()
                send_telegram_msg(chat_id, t('multiple_transactions_created').format(count=created_count))
                try:
                    streak_msg = _check_streak(workspace.id, db, t)
                    if streak_msg:
                        send_telegram_msg(chat_id, streak_msg)
                except Exception:
                    pass
                return {'status': 'success'}

            db.flush()
            # Construir mensagem com botões por linha (✅/❌) + Confirmar tudo / Cancelar tudo
            message_text, reply_markup = _build_batch_message_and_keyboard(str(chat_id), batch_id_hex, db, t)
            if message_text and reply_markup:
                send_telegram_msg(chat_id, message_text, reply_markup)
            db.commit()
            return {'status': 'success'}
            
        # Processar transação única
        amount_cents = int(parsed['amount'] * 100)
        transaction_date = parsed.get('transaction_date') or date.today()
        suggested_category_name = parsed.get('suggested_category_name')
        if parsed.get('type') == 'expense':
            amount_cents = -abs(amount_cents)
        else:
            amount_cents = abs(amount_cents)

        # Deduplicação: já existe pendente com mesma descrição+valor+tipo?
        if not parsed.get('multiple') and not user.telegram_auto_confirm and parsed.get('category_id') and not suggested_category_name:
            cache_key = _description_cache_key(parsed.get('description') or "")
            existing = db.query(models.TelegramPendingTransaction).filter(
                models.TelegramPendingTransaction.chat_id == str(chat_id),
                models.TelegramPendingTransaction.workspace_id == workspace.id,
                models.TelegramPendingTransaction.amount_cents == amount_cents,
                models.TelegramPendingTransaction.batch_id.is_(None),
            ).all()
            for ex in existing:
                if _description_cache_key(ex.description) == cache_key:
                    send_telegram_msg(chat_id, t('pending_duplicate'))
                    return {'status': 'duplicate_pending'}
        
        # Se a IA sugeriu uma categoria que não existe: perguntar se quer criar
        if suggested_category_name and not parsed.get('category_id'):
            pending = models.TelegramPendingTransaction(
                chat_id=str(chat_id),
                workspace_id=workspace.id,
                category_id=None,
                amount_cents=amount_cents,
                description=(parsed['description'] or "")[:255],
                inference_source=parsed.get('inference_source'),
                decision_reason=parsed.get('decision_reason'),
                needs_review=True,
                transaction_date=transaction_date,
                suggested_category_name=suggested_category_name[:100],
            )
            db.add(pending)
            db.commit()
            message_text = t('create_category_prompt').format(name=suggested_category_name)
            pending_id_hex = pending.id.hex[:16]
            reply_markup = {
                "inline_keyboard": [[
                    {"text": t('button_create_category'), "callback_data": f"create_cat_{pending_id_hex}"},
                    {"text": t('button_skip_category'), "callback_data": f"skip_cat_{pending_id_hex}"}
                ]]
            }
            send_telegram_msg(chat_id, message_text, reply_markup)
            logger.info("Pendente com suggested_category_name criado; aguardando user criar ou cancelar")
            return {'status': 'success'}

        category = db.query(models.Category).filter(
            models.Category.id == parsed['category_id']
        ).first()
        category_name = category.name if category else "Outros"

        # Lógica especial para categorias de vault (investimento/emergência)
        is_vault_category = category and category.vault_type != 'none'
        is_vault_withdrawal = parsed.get('is_vault_withdrawal', False) if is_vault_category else False
        if is_vault_category:
            amount_cents = -abs(amount_cents) if is_vault_withdrawal else abs(amount_cents)

        if user.telegram_auto_confirm:
            logger.info("Modo auto_confirm ativo - criando transacao diretamente")
            transaction = models.Transaction(
                workspace_id=workspace.id,
                category_id=parsed['category_id'],
                amount_cents=amount_cents,
                description=parsed['description'],
                inference_source=parsed.get('inference_source'),
                decision_reason=parsed.get('decision_reason'),
                needs_review=parsed.get('needs_review', False),
                transaction_date=transaction_date,
            )
            db.add(transaction)
            db.flush()
            logger.info("Transacao criada com ID: %s, workspace_id: %s, amount_cents: %s", transaction.id, workspace.id, amount_cents)
            db.commit()
            logger.info("Transacao commitada com sucesso (auto_confirm)")
            tipo_emoji = "" if amount_cents < 0 else "💰"
            tipo_texto = t('type_expense') if amount_cents < 0 else t('type_income')
            origin_line = _origin_line(parsed.get('inference_source'), t)
            send_telegram_msg(chat_id, t('transaction_registered').format(
                description=parsed['description'],
                emoji=tipo_emoji,
                amount=abs(parsed['amount']),
                category=category_name,
                type=tipo_texto,
                origin_line=origin_line,
                date_line=_date_line(transaction_date, t),
            ))
            # Budget alert, streak, insight after auto-confirm
            try:
                alert = _check_budget_alerts(workspace.id, parsed['category_id'], db, t)
                if alert:
                    send_telegram_msg(chat_id, alert)
                streak_msg = _check_streak(workspace.id, db, t)
                if streak_msg:
                    send_telegram_msg(chat_id, streak_msg)
                insight_msg = _generate_insight(workspace.id, parsed['category_id'], db, t)
                if insight_msg:
                    send_telegram_msg(chat_id, insight_msg)
                month_cmp = _check_month_comparison(workspace.id, db, t, user.language or 'pt')
                if month_cmp:
                    send_telegram_msg(chat_id, month_cmp)
            except Exception as extra_err:
                logger.warning("Erro extras post-autoconfirm: %s", extra_err)
        else:
            pending = models.TelegramPendingTransaction(
                chat_id=str(chat_id),
                workspace_id=workspace.id,
                category_id=parsed['category_id'],
                amount_cents=amount_cents,
                description=parsed['description'],
                inference_source=parsed.get('inference_source'),
                decision_reason=parsed.get('decision_reason'),
                needs_review=parsed.get('needs_review', False),
                transaction_date=transaction_date,
            )
            db.add(pending)
            db.commit()
            tipo_emoji = "" if amount_cents < 0 else "💰"
            tipo_texto = t('type_expense') if amount_cents < 0 else t('type_income')
            origin_line = _origin_line(parsed.get('inference_source'), t)
            message_text = t('transaction_pending').format(
                description=parsed['description'],
                emoji=tipo_emoji,
                amount=abs(parsed['amount']),
                category=category_name,
                type=tipo_texto,
                origin_line=origin_line,
                date_line=_date_line(transaction_date, t),
            )
            pending_id_hex = pending.id.hex[:16]
            reply_markup = {
                "inline_keyboard": [[
                    {"text": t('button_confirm'), "callback_data": f"confirm_{pending_id_hex}"},
                    {"text": t('button_cancel'), "callback_data": f"cancel_{pending_id_hex}"},
                ]]
            }
            send_telegram_msg(chat_id, message_text, reply_markup)

        logger.info("Transação processada com sucesso")
        return {'status': 'success'}
        
    except Exception as e:
        logger.error(f"Erro Telegram: {str(e)}", exc_info=True)
        import traceback
        logger.error(f"Traceback completo: {traceback.format_exc()}")
        try:
            chat_id = (data.get('message') or {}).get('chat', {}).get('id') or (data.get('callback_query') or {}).get('message', {}).get('chat', {}).get('id') if data else None
            if chat_id and db:
                user_temp = db.query(models.User).filter(models.User.phone_number == str(chat_id)).first()
                lang = (user_temp.language if user_temp and user_temp.language else None) or 'pt'
                t_err = get_telegram_t(lang)
                send_telegram_msg(chat_id, t_err('generic_error'))
        except Exception:
            pass
        return {'status': 'error'}
